import streamlit as st
import requests
import json
import pandas as pd
from habanero import Crossref
from crossref_commons.retrieval import get_publication_as_json
import time
from typing import List, Dict, Any, Tuple, Set
from datetime import datetime
from collections import Counter
import re
import numpy as np
from functools import lru_cache
import os
import tempfile
import zipfile
import shutil
from tqdm import tqdm
from contextlib import redirect_stdout
from io import StringIO, BytesIO
from ratelimit import limits, sleep_and_retry
from tenacity import retry, stop_after_attempt, wait_exponential, retry_if_exception_type
import logging
from bs4 import BeautifulSoup
import io
import csv
import nltk
from nltk.corpus import stopwords
from nltk.stem import PorterStemmer
import base64
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import networkx as nx
from itertools import combinations
import math

try:
    nltk.data.find('corpora/stopwords')
except LookupError:
    nltk.download('stopwords')

class Config:
    REQUEST_TIMEOUT = 30
    MAX_RETRIES = 5
    DELAY_BETWEEN_REQUESTS = 0.31
    RETRY_DELAY = 1

class PerformanceMonitor:
    def __init__(self):
        self.start_time = None
        self.request_count = 0

    def start(self):
        self.start_time = datetime.now()

    def increment_request(self):
        self.request_count += 1

    def get_stats(self):
        if self.start_time:
            elapsed = (datetime.now() - self.start_time).total_seconds()
            return {
                'total_requests': self.request_count,
                'elapsed_seconds': elapsed,
                'elapsed_minutes': elapsed / 60,
                'requests_per_second': self.request_count / elapsed if elapsed > 0 else 0
            }
        return {}

class EthicsDetector:
    """Детектор неэтичных практик цитирования"""
    
    def __init__(self):
        self.suspicious_patterns = []
        
    def detect_citation_burst(self, source_articles_df: pd.DataFrame, citations_df: pd.DataFrame, 
                            threshold_ratio: float = 10.0) -> List[Dict]:
        """
        Обнаружение внезапных всплесков цитирований
        threshold_ratio: минимальное отношение цитирований к возрасту статьи для подозрения
        """
        suspicious = []
        
        for _, article in source_articles_df.iterrows():
            try:
                citation_count = max(article.get('citation_count_openalex', 0), 
                                   article.get('citation_count_crossref', 0))
                years_since_pub = article.get('years_since_publication', 1)
                
                if years_since_pub <= 0:
                    years_since_pub = 1
                
                citation_ratio = citation_count / years_since_pub
                
                if citation_ratio > threshold_ratio:
                    suspicious.append({
                        'type': 'CITATION_BURST',
                        'severity': 'HIGH',
                        'article_doi': article.get('doi', 'Unknown'),
                        'citation_count': citation_count,
                        'years_since_publication': years_since_pub,
                        'citation_ratio': round(citation_ratio, 2),
                        'threshold': threshold_ratio,
                        'description': f"Статья имеет {citation_count} цитирований за {years_since_pub} год(а) - соотношение {citation_ratio:.1f} цитирований/год"
                    })
            except Exception as e:
                continue
                
        return suspicious
    
    def detect_self_citation(self, source_articles_df: pd.DataFrame, references_df: pd.DataFrame, 
                           threshold_percentage: float = 40.0) -> List[Dict]:
        """
        Обнаружение чрезмерного самоцитирования
        threshold_percentage: процент самоцитирований для подозрения
        """
        suspicious = []
        
        # Создаем словарь авторов для каждой статьи
        article_authors = {}
        for _, article in source_articles_df.iterrows():
            authors = self._extract_author_names(article.get('authors', ''))
            article_authors[article.get('doi')] = authors
        
        # Анализируем ссылки для каждой статьи
        for source_doi, authors in article_authors.items():
            if not authors:
                continue
                
            article_refs = references_df[references_df['source_doi'] == source_doi]
            total_refs = len(article_refs)
            
            if total_refs == 0:
                continue
                
            self_citation_count = 0
            
            for _, ref in article_refs.iterrows():
                ref_authors = self._extract_author_names(ref.get('authors', ''))
                if self._has_author_overlap(authors, ref_authors):
                    self_citation_count += 1
            
            self_citation_percentage = (self_citation_count / total_refs) * 100
            
            if self_citation_percentage > threshold_percentage:
                suspicious.append({
                    'type': 'EXCESSIVE_SELF_CITATION',
                    'severity': 'MEDIUM',
                    'article_doi': source_doi,
                    'self_citation_count': self_citation_count,
                    'total_references': total_refs,
                    'self_citation_percentage': round(self_citation_percentage, 1),
                    'threshold': threshold_percentage,
                    'description': f"{self_citation_percentage:.1f}% ссылок ({self_citation_count}/{total_refs}) являются самоцитированием"
                })
                
        return suspicious
    
    def detect_citation_cartels(self, references_df: pd.DataFrame, 
                              affiliation_threshold: float = 0.7,
                              citation_threshold: float = 0.6) -> List[Dict]:
        """
        Обнаружение цитатных колец (групп авторов/организаций, цитирующих друг друга)
        """
        suspicious = []
        
        try:
            # Строим граф цитирований между аффилиациями
            affiliation_citation_graph = nx.Graph()
            affiliation_citations = {}
            
            for _, ref in references_df.iterrows():
                source_affiliations = self._extract_affiliations(ref.get('affiliations', ''))
                citing_affiliations = self._extract_affiliations(ref.get('affiliations', ''))
                
                for source_affil in source_affiliations:
                    for citing_affil in citing_affiliations:
                        if source_affil != citing_affil and source_affil != 'Unknown' and citing_affil != 'Unknown':
                            key = (source_affil, citing_affil)
                            affiliation_citations[key] = affiliation_citations.get(key, 0) + 1
                            
                            if affiliation_citation_graph.has_edge(source_affil, citing_affil):
                                affiliation_citation_graph[source_affil][citing_affil]['weight'] += 1
                            else:
                                affiliation_citation_graph.add_edge(source_affil, citing_affil, weight=1)
            
            # Анализируем плотность цитирований внутри групп
            if len(affiliation_citation_graph.nodes()) > 0:
                # Ищем клики и плотные подграфы
                cliques = list(nx.find_cliques(affiliation_citation_graph))
                
                for clique in cliques:
                    if len(clique) >= 3:  # Минимум 3 организации для кольца
                        internal_citations = 0
                        total_possible_citations = len(clique) * (len(clique) - 1)
                        
                        for org1, org2 in combinations(clique, 2):
                            if affiliation_citation_graph.has_edge(org1, org2):
                                internal_citations += affiliation_citation_graph[org1][org2]['weight']
                        
                        if total_possible_citations > 0:
                            density = internal_citations / total_possible_citations
                            
                            if density > citation_threshold:
                                suspicious.append({
                                    'type': 'CITATION_CARTEL',
                                    'severity': 'HIGH',
                                    'organizations': clique,
                                    'internal_citations': internal_citations,
                                    'total_possible_citations': total_possible_citations,
                                    'density': round(density, 3),
                                    'threshold': citation_threshold,
                                    'description': f"Обнаружено цитатное кольцо из {len(clique)} организаций с плотностью цитирований {density:.1%}"
                                })
                                
        except Exception as e:
            # В случае ошибки просто возвращаем пустой список
            pass
            
        return suspicious
    
    def detect_newborn_citation(self, references_df: pd.DataFrame, 
                              max_days_old: int = 30) -> List[Dict]:
        """
        Обнаружение цитирования очень новых статей
        max_days_old: максимальный возраст статьи в днях для подозрения
        """
        suspicious = []
        
        current_year = datetime.now().year
        current_month = datetime.now().month
        
        for _, ref in references_df.iterrows():
            try:
                ref_year = ref.get('year', 'Unknown')
                if ref_year == 'Unknown':
                    continue
                    
                ref_year = int(ref_year)
                ref_age_years = current_year - ref_year
                
                # Если статья текущего года, считаем примерный возраст в месяцах
                if ref_year == current_year:
                    # Упрощенная логика - считаем, что статья не должна быть младше 1 месяца
                    ref_age_days = 30  # Консервативная оценка
                    
                    if ref_age_days < max_days_old:
                        source_doi = ref.get('source_doi', 'Unknown')
                        suspicious.append({
                            'type': 'NEWBORN_CITATION',
                            'severity': 'LOW',
                            'source_doi': source_doi,
                            'cited_doi': ref.get('doi', 'Unknown'),
                            'cited_year': ref_year,
                            'estimated_age_days': ref_age_days,
                            'threshold': max_days_old,
                            'description': f"Статья цитирует очень новую публикацию ({ref_year} года), возраст ~{ref_age_days} дней"
                        })
                        
            except (ValueError, TypeError):
                continue
                
        return suspicious
    
    def detect_mass_citation_by_author(self, references_df: pd.DataFrame,
                                     threshold_count: int = 50) -> List[Dict]:
        """
        Обнаружение массового цитирования одним автором
        threshold_count: минимальное количество цитирований для подозрения
        """
        suspicious = []
        
        try:
            # Анализируем частоту цитирования по авторам
            author_citation_count = {}
            
            for _, ref in references_df.iterrows():
                authors = self._extract_author_names(ref.get('authors', ''))
                for author in authors:
                    author_citation_count[author] = author_citation_count.get(author, 0) + 1
            
            # Проверяем авторов с подозрительно высокой активностью
            for author, count in author_citation_count.items():
                if count > threshold_count and author != 'Unknown':
                    suspicious.append({
                        'type': 'MASS_CITATION_BY_AUTHOR',
                        'severity': 'MEDIUM',
                        'author': author,
                        'citation_count': count,
                        'threshold': threshold_count,
                        'description': f"Автор {author} имеет {count} цитирований, что превышает порог в {threshold_count}"
                    })
                    
        except Exception as e:
            pass
            
        return suspicious
    
    def detect_citation_snowball(self, references_df: pd.DataFrame,
                               similarity_threshold: float = 0.8) -> List[Dict]:
        """
        Обнаружение "цитатного снежного кома" - статей с очень однородными цитированиями
        similarity_threshold: порог схожести паттернов цитирования
        """
        suspicious = []
        
        try:
            # Группируем ссылки по исходным статьям
            source_articles = references_df['source_doi'].unique()
            
            if len(source_articles) < 2:
                return suspicious
            
            # Создаем матрицу цитирований
            citation_patterns = {}
            for source_doi in source_articles:
                refs = references_df[references_df['source_doi'] == source_doi]
                cited_dois = set(refs['doi'].dropna())
                citation_patterns[source_doi] = cited_dois
            
            # Сравниваем паттерны цитирования
            for doi1, doi2 in combinations(source_articles, 2):
                pattern1 = citation_patterns[doi1]
                pattern2 = citation_patterns[doi2]
                
                if not pattern1 or not pattern2:
                    continue
                
                intersection = len(pattern1.intersection(pattern2))
                union = len(pattern1.union(pattern2))
                
                if union > 0:
                    similarity = intersection / union
                    
                    if similarity > similarity_threshold:
                        suspicious.append({
                            'type': 'CITATION_SNOWBALL',
                            'severity': 'MEDIUM',
                            'article_doi_1': doi1,
                            'article_doi_2': doi2,
                            'common_citations': intersection,
                            'total_unique_citations': union,
                            'similarity': round(similarity, 3),
                            'threshold': similarity_threshold,
                            'description': f"Статьи {doi1} и {doi2} имеют {similarity:.1%} схожих цитирований ({intersection} общих из {union} уникальных)"
                        })
                        
        except Exception as e:
            pass
            
        return suspicious

    def detect_citation_templating(self, references_df: pd.DataFrame,
                                 template_size: int = 5,
                                 occurrence_threshold: int = 10) -> List[Dict]:
        """
        Обнаружение шаблонного цитирования (одинаковые наборы цитирований)
        template_size: размер шаблона для поиска
        occurrence_threshold: минимальное количество повторений
        """
        suspicious = []
    
        try:
            import random  # Локальный импорт для сэмплинга (если не добавлен глобально)
        
            # Создаем наборы цитирований для каждой статьи
            article_citation_sets = {}
            for source_doi in references_df['source_doi'].unique():
                refs = references_df[references_df['source_doi'] == source_doi]
                cited_dois = list(refs['doi'].dropna())  # Преобразуем в list для sample
                if len(cited_dois) >= template_size:
                    # Сэмплируем, если слишком много (избежать memory boom)
                    if len(cited_dois) > 50:
                        cited_dois = random.sample(cited_dois, 50)  # Max 50 для безопасности
                        st.warning(f"Сэмплировано 50 DOI для {source_doi} (было {len(refs['doi'].dropna())})")
                    article_citation_sets[source_doi] = set(cited_dois)
        
            # Ищем часто встречающиеся комбинации
            citation_combinations = {}
        
            for doi, citation_set in article_citation_sets.items():
                effective_size = min(template_size, len(citation_set))
                # Генерируем только если мало (лимит для безопасности)
                if effective_size > 15:  # C(15,5)=3003, безопасно; больше — skip
                    st.warning(f"Пропущен большой набор для {doi}: {len(citation_set)} DOI слишком много для шаблона {effective_size}")
                    continue
            
            for combo in combinations(citation_set, effective_size):
                combo_key = frozenset(combo)
                citation_combinations[combo_key] = citation_combinations.get(combo_key, []) + [doi]
        
            # Проверяем комбинации, встречающиеся часто
            for combo, articles in citation_combinations.items():
                if len(articles) >= occurrence_threshold:
                    suspicious.append({
                        'type': 'CITATION_TEMPLATING',
                        'severity': 'HIGH',
                        'template_size': len(combo),
                        'articles_using_template': len(articles),
                        'template_dois': list(combo)[:10],  # Показываем первые 10 DOI
                        'affected_articles': articles[:10],  # Показываем первые 10 статей
                        'threshold': occurrence_threshold,
                        'description': f"Шаблон из {len(combo)} цитирований используется в {len(articles)} статьях"
                    })
                
        except Exception as e:
            st.error(f"Ошибка в detect_citation_templating: {e}")  # Лог вместо pass
            pass
        
        return suspicious
    
    def run_complete_analysis(self, references_df: pd.DataFrame, source_articles_df: pd.DataFrame) -> Dict:
        """Запуск полного анализа на неэтичные практики"""
        
        results = {
            'citation_burst': self.detect_citation_burst(source_articles_df, references_df),
            'self_citation': self.detect_self_citation(source_articles_df, references_df),
            'citation_cartels': self.detect_citation_cartels(references_df),
            'newborn_citation': self.detect_newborn_citation(references_df),
            'mass_citation_author': self.detect_mass_citation_by_author(references_df),
            'citation_snowball': self.detect_citation_snowball(references_df),
            'citation_templating': self.detect_citation_templating(references_df)
        }
        
        # Сводная статистика
        total_findings = sum(len(findings) for findings in results.values())
        severity_counts = {
            'HIGH': 0,
            'MEDIUM': 0, 
            'LOW': 0
        }
        
        for findings in results.values():
            for finding in findings:
                severity_counts[finding.get('severity', 'LOW')] += 1
        
        results['summary'] = {
            'total_findings': total_findings,
            'severity_counts': severity_counts,
            'analysis_timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        
        return results
    
    def _extract_author_names(self, authors_str: str) -> List[str]:
        """Извлекает список авторов из строки"""
        if not authors_str or authors_str in ['Unknown', 'Error']:
            return []
        
        authors = []
        for author in authors_str.split(','):
            author = author.strip()
            if author and author not in ['Unknown', 'Error']:
                # Берем только фамилию (первое слово)
                surname = author.split()[0] if author.split() else author
                authors.append(surname)
        
        return authors
    
    def _has_author_overlap(self, authors1: List[str], authors2: List[str]) -> bool:
        """Проверяет пересечение авторов"""
        if not authors1 or not authors2:
            return False
        
        return len(set(authors1).intersection(set(authors2))) > 0
    
    def _extract_affiliations(self, affiliations_str: str) -> List[str]:
        """Извлекает список аффилиаций из строки"""
        if not affiliations_str or affiliations_str in ['Unknown', 'Error']:
            return []
        
        affiliations = []
        for affil in affiliations_str.split(';'):
            affil = affil.strip()
            if affil and affil not in ['Unknown', 'Error']:
                affiliations.append(affil)
        
        return affiliations

class FastAffiliationProcessor:
    """Быстрый процессор аффилиаций с группировкой похожих организаций"""

    def __init__(self):
        self.common_keywords = {
            'university', 'college', 'institute', 'school', 'department', 'faculty',
            'laboratory', 'center', 'centre', 'academy', 'universität', 'universitat',
            'université', 'universite', 'polytechnic', 'technical', 'technology',
            'research', 'science', 'sciences', 'studies', 'medical', 'hospital',
            'clinic', 'foundation', 'corporation', 'company', 'inc', 'ltd', 'corp'
        }
        self.organization_cache = {}
        self.country_keywords = {
            'usa', 'united states', 'us', 'u.s.', 'u.s.a.', 'america',
            'uk', 'united kingdom', 'great britain', 'england', 'scotland', 'wales',
            'germany', 'deutschland', 'france', 'french', 'italy', 'italian',
            'spain', 'spanish', 'china', 'chinese', 'japan', 'japanese',
            'russia', 'russian', 'india', 'indian', 'brazil', 'brazilian',
            'canada', 'canadian', 'australia', 'australian', 'korea', 'korean'
        }

    def extract_main_organization_fast(self, affiliation: str) -> str:
        """Быстрое извлечение основной организации из полной аффилиации"""
        if not affiliation or affiliation in ['Unknown', 'Error', '']:
            return "Unknown"

        # Кеширование результатов
        if affiliation in self.organization_cache:
            return self.organization_cache[affiliation]

        # Очистка текста
        clean_affiliation = affiliation.strip()

        # Удаляем email-адреса
        clean_affiliation = re.sub(r'\S+@\S+', '', clean_affiliation)

        # Удаляем почтовые индексы и адреса
        clean_affiliation = re.sub(r'\d{5,}(?:-\d{4})?', '', clean_affiliation)
        clean_affiliation = re.sub(r'p\.?o\.? box \d+', '', clean_affiliation, flags=re.IGNORECASE)
        clean_affiliation = re.sub(r'\b\d+\s+[a-zA-Z]+\s+[a-zA-Z]+\b', '', clean_affiliation)  # адреса типа "123 Main Street"

        # Разделяем по запятым, точкам с запятой и другим разделителям
        parts = re.split(r'[,;]', clean_affiliation)

        # Ищем часть с основной организацией
        main_org_candidates = []

        for part in parts:
            part = part.strip()
            if not part or len(part) < 5:  # Слишком короткие части пропускаем
                continue

            part_lower = part.lower()

            # Проверяем, содержит ли часть ключевые слова организации
            has_org_keyword = any(keyword in part_lower for keyword in self.common_keywords)

            # Проверяем, не содержит ли страну (это обычно не основная организация)
            has_country = any(country in part_lower for country in self.country_keywords)

            if has_org_keyword and not has_country:
                main_org_candidates.append(part)

        # Выбираем лучшего кандидата
        if main_org_candidates:
            # Предпочитаем более длинные названия (обычно это полное название организации)
            main_org_candidates.sort(key=len, reverse=True)
            main_org = main_org_candidates[0]
        else:
            # Если не нашли по ключевым словам, берем первую значимую часть
            for part in parts:
                part = part.strip()
                if len(part) > 10 and not any(country in part.lower() for country in self.country_keywords):
                    main_org = part
                    break
            else:
                # Если все еще не нашли, берем первую непустую часть
                for part in parts:
                    part = part.strip()
                    if part:
                        main_org = part
                        break
                else:
                    main_org = clean_affiliation

        # Очищаем результат
        main_org = re.sub(r'\s+', ' ', main_org).strip()
        main_org = re.sub(r'^[^a-zA-Z0-9]*|[^a-zA-Z0-9]*$', '', main_org)

        result = main_org if main_org else "Unknown"
        self.organization_cache[affiliation] = result
        return result

    def normalize_organization_name(self, org_name: str) -> str:
        """Нормализует название организации для группировки"""
        if not org_name or org_name == "Unknown":
            return org_name

        # Приводим к нижнему регистру
        normalized = org_name.lower()

        # Удаляем общие префиксы и суффиксы
        remove_patterns = [
            r'^the\s+', r'\s+the$',
            r'\bdept\.?\s+of\b', r'\bdepartment\s+of\b',
            r'\bfaculty\s+of\b', r'\bschool\s+of\b',
            r'\binstitute\s+of\b', r'\binstitution\s+of\b',
            r'\bcollege\s+of\b', r'\bacademy\s+of\b',
            r'\blaboratory\b', r'\blab\b',
            r'\bcenter\b', r'\bcentre\b',
            r'\bdivision\b', r'\bgroup\b',
            r'\binc\.?$', r'\bltd\.?$', r'\bcorp\.?$', r'\bco\.?$',
            r'\bllc\.?$', r'\bgmbh\.?$'
        ]

        for pattern in remove_patterns:
            normalized = re.sub(pattern, '', normalized)

        # Удаляем лишние пробелы и символы
        normalized = re.sub(r'\s+', ' ', normalized).strip()
        normalized = re.sub(r'[^\w\s&]', '', normalized)

        return normalized.strip()

    def group_similar_organizations(self, organizations: List[str]) -> Dict[str, List[str]]:
        """Группирует похожие организации"""
        if not organizations:
            return {}

        # Создаем нормализованные версии
        normalized_map = {}
        for org in organizations:
            if org != "Unknown":
                normalized = self.normalize_organization_name(org)
                if normalized:
                    if normalized not in normalized_map:
                        normalized_map[normalized] = []
                    normalized_map[normalized].append(org)

        # Объединяем очень похожие группы
        final_groups = {}
        normalized_keys = list(normalized_map.keys())

        for i, key1 in enumerate(normalized_keys):
            if key1 not in final_groups:
                final_groups[key1] = []

            # Добавляем все организации для этой группы
            final_groups[key1].extend(normalized_map[key1])

            # Ищем похожие группы для объединения
            for j, key2 in enumerate(normalized_keys[i+1:], i+1):
                if self.are_organizations_similar(key1, key2):
                    if key2 in normalized_map:
                        final_groups[key1].extend(normalized_map[key2])
                    # Помечаем для удаления
                    if key2 in final_groups:
                        del final_groups[key2]

        return final_groups

    def are_organizations_similar(self, org1: str, org2: str) -> bool:
        """Проверяет, являются ли две организации похожими"""
        if not org1 or not org2:
            return False

        org1_lower = org1.lower()
        org2_lower = org2.lower()

        # Простая проверка на вхождение
        if org1_lower in org2_lower or org2_lower in org1_lower:
            return True

        # Разбиваем на слова и проверяем пересечение
        words1 = set(org1_lower.split())
        words2 = set(org2_lower.split())

        if not words1 or not words2:
            return False

        # Удаляем стоп-слова
        stop_words = {'the', 'and', 'of', 'for', 'in', 'on', 'at', 'to', 'by'}
        words1 = words1 - stop_words
        words2 = words2 - stop_words

        if not words1 or not words2:
            return False

        # Вычисляем коэффициент Жаккара
        intersection = words1.intersection(words2)
        union = words1.union(words2)

        similarity = len(intersection) / len(union) if union else 0

        return similarity > 0.6  # Порог схожести

    def process_affiliations_list_fast(self, affiliations: List[str]) -> Tuple[Dict[str, int], Dict[str, List[str]]]:
        """Быстрая обработка списка аффилиаций с группировкой"""
        if not affiliations:
            return {}, {}

        # Извлекаем основные организации
        main_organizations = []
        for aff in affiliations:
            if aff and aff not in ['Unknown', 'Error']:
                main_org = self.extract_main_organization_fast(aff)
                if main_org and main_org != "Unknown":
                    main_organizations.append(main_org)

        if not main_organizations:
            return {}, {}

        # Группируем похожие организации
        grouped_organizations = self.group_similar_organizations(main_organizations)

        # Выбираем представителя для каждой группы (самое частое или самое длинное название)
        group_representatives = {}
        for normalized_name, org_list in grouped_organizations.items():
            if org_list:
                # Выбираем самое длинное название как представителя (обычно это полное название организации)
                representative = max(org_list, key=len)
                group_representatives[representative] = org_list

        # Считаем частоты для представителей
        frequency_count = {}
        for representative, org_list in group_representatives.items():
            frequency_count[representative] = len(org_list)

        return frequency_count, group_representatives

class AltmetricProcessor:
    """Процессор для сбора альтметрических данных"""

    def __init__(self):
        self.altmetric_cache = {}

    def clean_doi(self, doi: str) -> str:
        """Очищает DOI от лишних символов, префиксов и пробелов."""
        if not doi or doi in ['Unknown', 'Error', '']:
            return None

        doi = doi.strip().lower()
        doi = re.sub(r'^(doi:)?\s*', '', doi)  # Удаляет "doi:" и пробелы
        doi = re.sub(r'\s+', '', doi)  # Удаляет лишние пробелы
        if re.match(r'^10\.\d{4,9}/[-._;()/:a-zA-Z0-9]+$', doi):
            return doi
        return None

    @retry(stop=stop_after_attempt(Config.MAX_RETRIES),
           wait=wait_exponential(multiplier=1, min=Config.RETRY_DELAY, max=10),
           retry=retry_if_exception_type((requests.exceptions.RequestException,
                                        requests.exceptions.Timeout,
                                        requests.exceptions.ConnectionError)))
    @sleep_and_retry
    @limits(calls=15, period=1)
    def get_altmetric_data(self, doi: str) -> Dict:
        """Получает данные из бесплатного API Altmetric по DOI."""
        clean_doi = self.clean_doi(doi)
        if not clean_doi:
            return None

        if clean_doi in self.altmetric_cache:
            return self.altmetric_cache[clean_doi]

        url = f"https://api.altmetric.com/v1/doi/{clean_doi}"
        try:
            response = requests.get(url, timeout=Config.REQUEST_TIMEOUT)
            if response.status_code == 200:
                data = response.json()
                self.altmetric_cache[clean_doi] = data
                return data
            else:
                self.altmetric_cache[clean_doi] = None
                return None
        except requests.exceptions.RequestException:
            self.altmetric_cache[clean_doi] = None
            return None

    def get_altmetric_metrics(self, doi: str) -> Dict[str, Any]:
        """Извлекает ключевые альтметрические показатели для DOI"""
        data = self.get_altmetric_data(doi)

        if not data:
            return {
                'altmetric_score': 0,
                'cited_by_posts_count': 0,
                'cited_by_tweeters_count': 0,
                'cited_by_feeds_count': 0,
                'cited_by_accounts_count': 0
            }

        return {
            'altmetric_score': data.get('score', 0),
            'cited_by_posts_count': data.get('cited_by_posts_count', 0),
            'cited_by_tweeters_count': data.get('cited_by_tweeters_count', 0),
            'cited_by_feeds_count': data.get('cited_by_feeds_count', 0),
            'cited_by_accounts_count': data.get('cited_by_accounts_count', 0)
        }

class CitationAnalyzer:
    def __init__(self, rate_limit_calls=10, rate_limit_period=1):
        self.crossref_cache = {}
        self.openalex_cache = {}
        self.rate_limit_calls = rate_limit_calls
        self.rate_limit_period = rate_limit_period
        self.performance_monitor = PerformanceMonitor()
        self._unique_references_cache = {}
        self._unique_citations_cache = {}
        self.unique_ref_data_cache = {}
        self.unique_citation_data_cache = {}
        self.ltwa_map = None
        self.stop_words = set(stopwords.words('english'))
        self.stemmer = PorterStemmer()
        self.fast_affiliation_processor = FastAffiliationProcessor()
        self.altmetric_processor = AltmetricProcessor()
        self.ethics_detector = EthicsDetector()  # Новый детектор неэтичных практик
        self.scientific_stopwords = {
            'using', 'based', 'study', 'studies', 'research', 'analysis',
            'effect', 'effects', 'properties', 'property', 'development',
            'application', 'applications', 'method', 'methods', 'approach',
            'review', 'investigation', 'characterization', 'evaluation',
            'performance', 'behavior', 'structure', 'synthesis', 'design',
            'fabrication', 'preparation', 'processing', 'measurement',
            'model', 'models', 'system', 'systems', 'technology', 'material',
            'materials', 'sample', 'samples', 'device', 'devices', 'film',
            'films', 'layer', 'layers', 'surface', 'surfaces', 'interface',
            'interfaces', 'nanoparticle', 'nanoparticles', 'nanostructure',
            'nanostructures', 'composite', 'composites', 'coating', 'coatings'
        }
        self.scientific_stopwords_stemmed = {self.stemmer.stem(word) for word in self.scientific_stopwords}
        self.setup_logging()

    def setup_logging(self):
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(f'doi_analyzer_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log'),
                logging.StreamHandler()
            ]
        )
        self.logger = logging.getLogger(__name__)

    def validate_doi(self, doi: str) -> bool:
        """Проверяет валидность DOI с улучшенной обработкой"""
        if not doi or not isinstance(doi, str):
            return False

        doi = self.normalize_doi(doi)

        doi_pattern = r'^10\.\d{4,9}/[-._;()/:a-zA-Z0-9]+$'

        if not bool(re.match(doi_pattern, doi, re.IGNORECASE)):
            return False

        if len(doi) < 10:
            return False

        if re.search(r'[^\w\.\-_;()/:]', doi):
            return False

        return True

    def normalize_doi(self, doi: str) -> str:
        """Нормализует DOI, убирая префиксы и лишние символы"""
        if not doi or not isinstance(doi, str):
            return ""

        doi = doi.strip()

        prefixes = [
            'https://doi.org/', 'http://doi.org/', 'doi.org/',
            'doi:', 'DOI:', 'https://dx.doi.org/', 'http://dx.doi.org/',
        ]

        for prefix in prefixes:
            if doi.lower().startswith(prefix.lower()):
                doi = doi[len(prefix):]
                break

        doi = doi.split('?')[0].split('#')[0]
        doi = doi.strip()

        return doi.lower()

    def parse_doi_input(self, input_text: str, max_dois: int = 200) -> List[str]:
        """Парсит ввод DOI с улучшенной обработкой различных форматов"""
        if not input_text or not isinstance(input_text, str):
            st.error("Error: Input is empty or not a string")
            return []

        lines = input_text.strip().split('\n')
        dois = []

        for line in lines:
            line = line.strip()
            if not line:
                continue

            line = line.rstrip('.,;')
            doi_pattern = r'10\.\d{4,9}/[-._;()/:a-zA-Z0-9]+'
            found_dois = re.findall(doi_pattern, line, re.IGNORECASE)

            if found_dois:
                dois.extend(found_dois)
            else:
                if 'doi.org/' in line.lower():
                    doi_part = line.lower().split('doi.org/')[-1]
                    doi_part = doi_part.split('?')[0].split('#')[0].strip()
                    if self.validate_doi(doi_part):
                        dois.append(doi_part)
                elif line.lower().startswith('doi:'):
                    doi_part = line[4:].strip()
                    if self.validate_doi(doi_part):
                        dois.append(doi_part)
                elif self.validate_doi(line):
                    dois.append(line)

        cleaned_dois = []
        for doi in dois:
            normalized_doi = self.normalize_doi(doi)
            if self.validate_doi(normalized_doi):
                cleaned_dois.append(normalized_doi)

        unique_dois = []
        seen = set()
        for doi in cleaned_dois:
            if doi not in seen:
                seen.add(doi)
                unique_dois.append(doi)

        unique_dois = unique_dois[:max_dois]

        if not unique_dois:
            st.error("Error: No valid DOIs found in the input.")
            st.info("Please check the format. Valid examples:")
            st.code("  - 10.1234/abcd.1234")
            st.code("  - https://doi.org/10.1234/abcd.1234")
            st.code("  - doi:10.1234/abcd.1234")
        else:
            st.success(f"Found {len(unique_dois)} valid DOI(s)")
            if len(cleaned_dois) > len(unique_dois):
                st.info(f"Removed {len(cleaned_dois) - len(unique_dois)} duplicate DOI(s)")
            if len(cleaned_dois) > max_dois:
                st.info(f"Limited to first {max_dois} unique DOI(s) from {len(cleaned_dois)} found")

        return unique_dois

    @retry(stop=stop_after_attempt(Config.MAX_RETRIES),
           wait=wait_exponential(multiplier=1, min=Config.RETRY_DELAY, max=10),
           retry=retry_if_exception_type((requests.exceptions.RequestException,
                                        requests.exceptions.Timeout,
                                        requests.exceptions.ConnectionError)))
    @sleep_and_retry
    @limits(calls=15, period=1)
    def get_openalex_data(self, doi: str) -> Dict:
        """Получает данные из OpenAlex с повторными попытками"""
        if doi in self.openalex_cache:
            return self.openalex_cache[doi]
        try:
            openalex_url = f"https://api.openalex.org/works/https://doi.org/{doi}"
            response = requests.get(openalex_url, timeout=Config.REQUEST_TIMEOUT)
            self.performance_monitor.increment_request()
            if response.status_code == 404:
                self.openalex_cache[doi] = {}
                return {}
            response.raise_for_status()
            result = response.json()
            self.openalex_cache[doi] = result
            return result
        except requests.exceptions.RequestException as e:
            self.logger.warning(f"OpenAlex request failed for {doi}: {e}")
            if doi not in self.openalex_cache:
                self.openalex_cache[doi] = {}
            return {}
        except Exception as e:
            self.logger.error(f"Unexpected error in get_openalex_data for {doi}: {e}")
            if doi not in self.openalex_cache:
                self.openalex_cache[doi] = {}
            return {}

    @retry(stop=stop_after_attempt(Config.MAX_RETRIES),
           wait=wait_exponential(multiplier=1, min=Config.RETRY_DELAY, max=10),
           retry=retry_if_exception_type((requests.exceptions.RequestException,
                                        Exception)))
    @sleep_and_retry
    @limits(calls=15, period=1)
    def get_crossref_data(self, doi: str) -> Dict:
        """Получает данные из Crossref с повторными попытками и улучшенной обработкой аффилиаций"""
        if doi in self.crossref_cache:
            return self.crossref_cache[doi]
        try:
            cr = Crossref()
            result = cr.works(ids=doi)
            self.performance_monitor.increment_request()
            data = result['message']

            year = None
            for key in ['published-print', 'published-online', 'issued']:
                if key in data and 'date-parts' in data[key]:
                    date_parts = data[key]['date-parts'][0]
                    year = date_parts[0] if date_parts else None
                    break
            data['publication_year'] = year if year else 'Unknown'

            affiliations, countries = self.extract_affiliations_from_crossref(data)
            data['extracted_affiliations'] = affiliations
            data['extracted_countries'] = countries

            self.crossref_cache[doi] = data
            return data
        except Exception as e:
            self.logger.warning(f"Crossref request failed for {doi}: {e}")
            self.crossref_cache[doi] = {'publication_year': 'Unknown', 'extracted_affiliations': [], 'extracted_countries': []}
            return {'publication_year': 'Unknown', 'extracted_affiliations': [], 'extracted_countries': []}

    def extract_affiliations_from_crossref(self, crossref_data: Dict) -> tuple[List[str], List[str]]:
        """Извлекает аффилиации и страны из данных Crossref"""
        affiliations = set()
        countries = set()

        try:
            if 'author' in crossref_data:
                for author in crossref_data['author']:
                    if 'affiliation' in author:
                        for affil in author['affiliation']:
                            # Извлекаем основную организацию
                            if 'name' in affil:
                                affiliation_name = affil['name'].strip()
                                if affiliation_name and affiliation_name not in ['', 'None']:
                                    # Нормализуем аффилиацию
                                    main_org = self.fast_affiliation_processor.extract_main_organization_fast(affiliation_name)
                                    if main_org and main_org != "Unknown":
                                        affiliations.add(main_org)

                            # Извлекаем страну
                            country = self.extract_country_from_affiliation(affil)
                            if country:
                                countries.add(country)

            # Обрабатываем институциональные данные
            for field in ['institution', 'organization', 'department']:
                if field in crossref_data:
                    if isinstance(crossref_data[field], list):
                        for item in crossref_data[field]:
                            if isinstance(item, dict) and 'name' in item:
                                main_org = self.fast_affiliation_processor.extract_main_organization_fast(item['name'].strip())
                                if main_org and main_org != "Unknown":
                                    affiliations.add(main_org)
                            elif isinstance(item, str):
                                main_org = self.fast_affiliation_processor.extract_main_organization_fast(item.strip())
                                if main_org and main_org != "Unknown":
                                    affiliations.add(main_org)
                    elif isinstance(crossref_data[field], str):
                        main_org = self.fast_affiliation_processor.extract_main_organization_fast(crossref_data[field].strip())
                        if main_org and main_org != "Unknown":
                            affiliations.add(main_org)

        except Exception as e:
            self.logger.debug(f"Error extracting affiliations from Crossref: {e}")

        return list(affiliations), list(countries)

    def extract_country_from_affiliation(self, affiliation_data: Dict) -> str:
        """Извлекает страну из данных аффилиации Crossref"""
        try:
            if 'country' in affiliation_data and affiliation_data['country']:
                return affiliation_data['country'].upper().strip()

            if 'address' in affiliation_data and affiliation_data['address']:
                address = affiliation_data['address'].upper()
                country_codes = ['USA', 'US', 'UK', 'GB', 'DE', 'FR', 'CN', 'JP', 'RU', 'IN', 'BR', 'CA', 'AU', 'KR']
                for code in country_codes:
                    if code in address:
                        return code

            if 'name' in affiliation_data and affiliation_data['name']:
                name = affiliation_data['name'].upper()
                country_keywords = {
                    'UNITED STATES': 'US', 'USA': 'US', 'U.S.A': 'US', 'U.S.': 'US',
                    'UNITED KINGDOM': 'UK', 'UK': 'UK', 'GREAT BRITAIN': 'UK',
                    'GERMANY': 'DE', 'FRANCE': 'FR', 'CHINA': 'CN', 'JAPAN': 'JP',
                    'RUSSIA': 'RU', 'INDIA': 'IN', 'BRAZIL': 'BR', 'CANADA': 'CA',
                    'AUSTRALIA': 'AU', 'KOREA': 'KR', 'SOUTH KOREA': 'KR'
                }
                for keyword, code in country_keywords.items():
                    if keyword in name:
                        return code

        except Exception as e:
            self.logger.debug(f"Error extracting country from affiliation: {e}")

        return ""

    @sleep_and_retry
    @limits(calls=10, period=1)
    def quick_doi_search(self, title: str) -> str:
        """Quick DOI search by title"""
        if not title or title == 'Unknown':
            return None

        url = "https://api.crossref.org/works"
        params = {
            'query.title': title,
            'rows': 1,
            'select': 'DOI,title'
        }
        headers = {'User-Agent': 'CitationAnalyzer/1.0 (mailto:your@email.com)'}

        try:
            response = requests.get(url, params=params, headers=headers, timeout=Config.REQUEST_TIMEOUT)
            self.performance_monitor.increment_request()
            response.raise_for_status()
            data = response.json()

            if data['message']['items']:
                doi = data['message']['items'][0]['DOI']
                if self.validate_doi(doi):
                    return doi
            return None
        except:
            return None

    def safe_calculate_annual_citation_rate(self, citation_count, publication_year, current_year=None):
        """Безопасный расчет ежегодной цитируемости с обработкой ошибок"""
        try:
            if not isinstance(citation_count, (int, float)) or citation_count == 0:
                return 0.0

            years = self.calculate_years_since_publication(publication_year, current_year)

            if years is None or not isinstance(years, (int, float)) or years <= 0:
                return 0.0

            return round(citation_count / years, 2)
        except (TypeError, ZeroDivisionError, ValueError):
            return 0.0

    def calculate_years_since_publication(self, publication_year: Any, current_year: int = None) -> int:
        """Безопасный расчет лет с момента публикации"""
        try:
            if current_year is None:
                current_year = datetime.now().year

            if publication_year is None or publication_year == 'Unknown':
                return 1

            year_str = str(publication_year).strip()
            if not year_str or year_str == 'Unknown':
                return 1

            year_match = re.search(r'\b(19|20)\d{2}\b', year_str)
            if year_match:
                year = int(year_match.group())
            else:
                year = int(year_str)

            if 1900 < year <= current_year:
                return max(1, current_year - year)
            else:
                return 1
        except (ValueError, TypeError):
            return 1

    def get_combined_article_data(self, doi: str) -> Dict[str, Any]:
        """Get combined data from both Crossref and OpenAlex with improved affiliation processing and altmetrics"""
        try:
            crossref_data = self.get_crossref_data(doi)
            openalex_data = self.get_openalex_data(doi)
            altmetric_data = self.altmetric_processor.get_altmetric_metrics(doi)

            title = 'Unknown'
            if openalex_data and openalex_data.get('title'):
                title = openalex_data['title']
            elif crossref_data.get('title'):
                title_list = crossref_data['title']
                if title_list:
                    title = title_list[0]

            year = 'Unknown'
            publication_year = None

            if openalex_data and openalex_data.get('publication_year'):
                publication_year = openalex_data['publication_year']
                year = str(publication_year)
            elif crossref_data.get('publication_year') != 'Unknown':
                publication_year = crossref_data['publication_year']
                year = str(publication_year)

            authors = []
            authors_surnames = []
            authors_with_initials = []

            if openalex_data:
                for author in openalex_data.get('authorships', []):
                    name = author.get('author', {}).get('display_name', 'Unknown')
                    if name != 'Unknown':
                        authors.append(name)
                        surname_with_initial = self.extract_surname_with_initial(name)
                        authors_surnames.append(surname_with_initial)
                        authors_with_initials.append(surname_with_initial)

            if not authors and crossref_data.get('author'):
                for author in crossref_data['author']:
                    given = author.get('given', '')
                    family = author.get('family', '')
                    if given or family:
                        name = f"{given} {family}".strip()
                        authors.append(name)
                        surname_with_initial = self.extract_surname_with_initial(name)
                        authors_surnames.append(surname_with_initial)
                        authors_with_initials.append(surname_with_initial)

            authors_str = ', '.join(authors) if authors else 'Unknown'
            authors_surnames_str = ', '.join(authors_surnames) if authors_surnames else 'Unknown'
            authors_with_initials_str = ', '.join(authors_with_initials) if authors_with_initials else 'Unknown'
            author_count = len(authors) if authors else 0

            journal_info = self.get_journal_info_from_crossref(doi)

            _, crossref_citations, openalex_citations = self.get_citation_data(doi)

            # Улучшенная обработка аффилиаций
            affiliations, countries = self.get_enhanced_affiliations_and_countries(openalex_data, crossref_data)

            current_year = datetime.now().year
            years_since_pub = self.calculate_years_since_publication(publication_year, current_year)

            return {
                'doi': doi,
                'title': title,
                'year': year,
                'publication_year': publication_year,
                'authors': authors_str,
                'authors_surnames': authors_surnames_str,
                'authors_with_initials': authors_with_initials_str,
                'author_count': author_count,
                'journal_full_name': journal_info['full_name'],
                'journal_abbreviation': journal_info['abbreviation'],
                'publisher': journal_info['publisher'],
                'citation_count_crossref': crossref_citations,
                'citation_count_openalex': openalex_citations,
                'affiliations': '; '.join(affiliations),
                'countries': countries,
                'years_since_publication': years_since_pub,
                'altmetric_score': altmetric_data['altmetric_score'],
                'number_of_mentions': altmetric_data['cited_by_posts_count'],
                'x_mentions': altmetric_data['cited_by_tweeters_count'],
                'rss_blogs': altmetric_data['cited_by_feeds_count'],
                'unique_accounts': altmetric_data['cited_by_accounts_count']
            }
        except Exception as e:
            return {
                'doi': doi,
                'title': 'Error',
                'year': 'Unknown',
                'publication_year': None,
                'authors': 'Error',
                'authors_surnames': 'Error',
                'authors_with_initials': 'Error',
                'author_count': 0,
                'journal_full_name': 'Error',
                'journal_abbreviation': 'Error',
                'publisher': 'Error',
                'citation_count_crossref': 0,
                'citation_count_openalex': 0,
                'affiliations': 'Error',
                'countries': 'Error',
                'years_since_publication': 1,
                'altmetric_score': 0,
                'number_of_mentions': 0,
                'x_mentions': 0,
                'rss_blogs': 0,
                'unique_accounts': 0,
                'error': str(e)
            }

    def get_enhanced_affiliations_and_countries(self, openalex_data: Dict, crossref_data: Dict) -> tuple[List[str], str]:
        """Улучшенная обработка аффилиаций с группировкой"""
        try:
            # Получаем аффилиации из обоих источников
            openalex_affiliations, openalex_countries = self.get_affiliations_and_countries_from_openalex_data(openalex_data)
            crossref_affiliations, crossref_countries = self.get_affiliations_and_countries_from_crossref_data(crossref_data)

            # Объединяем аффилиации
            all_affiliations = []
            if openalex_affiliations and openalex_affiliations != ['Unknown']:
                all_affiliations.extend(openalex_affiliations)
            if crossref_affiliations:
                all_affiliations.extend(crossref_affiliations)

            # Обрабатываем аффилиации через быстрый процессор
            if all_affiliations:
                affiliation_frequencies, grouped_organizations = self.fast_affiliation_processor.process_affiliations_list_fast(all_affiliations)
                # Используем представителей групп как финальные аффилиации
                final_affiliations = list(affiliation_frequencies.keys())
            else:
                final_affiliations = ['Unknown']

            # Объединяем страны
            all_countries = set()
            if openalex_countries and openalex_countries != 'Unknown':
                countries_list = openalex_countries.split(';')
                all_countries.update([c.strip() for c in countries_list if c.strip()])
            if crossref_countries:
                all_countries.update(crossref_countries)

            final_countries = ';'.join(sorted(all_countries)) if all_countries else 'Unknown'

            return final_affiliations, final_countries

        except Exception as e:
            self.logger.error(f"Error in enhanced affiliations processing: {e}")
            return ['Unknown'], 'Unknown'

    def get_affiliations_and_countries_from_openalex_data(self, openalex_data: Dict) -> tuple[List[str], str]:
        """Извлекает аффилиации и страны из данных OpenAlex"""
        try:
            if not openalex_data:
                return ['Unknown'], 'Unknown'

            affiliations = set()
            countries = set()

            for authorship in openalex_data.get('authorships', []):
                for institution in authorship.get('institutions', []):
                    display_name = institution.get('display_name', '')
                    country_code = institution.get('country_code', '')

                    if display_name:
                        # Нормализуем название через наш процессор
                        main_org = self.fast_affiliation_processor.extract_main_organization_fast(display_name)
                        if main_org and main_org != "Unknown":
                            affiliations.add(main_org)

                    if country_code:
                        countries.add(country_code.upper())

            return list(affiliations) or ['Unknown'], ';'.join(sorted(countries)) if countries else 'Unknown'
        except Exception as e:
            self.logger.debug(f"Error extracting from OpenAlex: {e}")
            return ['Unknown'], 'Unknown'

    def get_affiliations_and_countries_from_crossref_data(self, crossref_data: Dict) -> tuple[List[str], List[str]]:
        """Извлекает аффилиации и страны из данных Crossref"""
        try:
            if not crossref_data:
                return [], []

            affiliations = crossref_data.get('extracted_affiliations', [])
            countries = crossref_data.get('extracted_countries', [])

            return affiliations, countries
        except Exception as e:
            self.logger.debug(f"Error extracting from Crossref: {e}")
            return [], []

    def get_citation_data(self, doi: str) -> tuple:
        try:
            crossref_data = self.get_crossref_data(doi)
            crossref_citations = crossref_data.get('is-referenced-by-count', 0)

            openalex_data = self.get_openalex_data(doi)
            openalex_citations = openalex_data.get('cited_by_count', 0)

            return doi, crossref_citations, openalex_citations
        except:
            return doi, 0, 0

    def get_journal_info_from_crossref(self, doi: str) -> Dict[str, Any]:
        try:
            data = self.get_crossref_data(doi)
            container_title = data.get('container-title', [])
            short_container_title = data.get('short-container-title', [])
            full_name = container_title[0] if container_title else (short_container_title[0] if short_container_title else 'Unknown')
            abbreviation = short_container_title[0] if short_container_title else (container_title[0] if container_title else 'Unknown')
            return {
                'full_name': full_name,
                'abbreviation': abbreviation,
                'publisher': data.get('publisher', 'Unknown'),
                'issn': data.get('ISSN', [None])[0]
            }
        except:
            return {
                'full_name': 'Unknown',
                'abbreviation': 'Unknown',
                'publisher': 'Unknown',
                'issn': None
            }

    @lru_cache(maxsize=1000)
    def extract_surname_with_initial(self, author_name: str) -> str:
        if not author_name or author_name in ['Unknown', 'Error']:
            return author_name
        clean_name = re.sub(r'[^\w\s\-\.]', ' ', author_name).strip()
        parts = clean_name.split()
        if not parts:
            return author_name
        surname = parts[-1]
        initial = parts[0][0].upper() if parts[0] else ''
        return f"{surname} {initial}." if initial else surname

    def get_references_from_crossref(self, doi: str) -> List[Dict[str, Any]]:
        try:
            article_data = get_publication_as_json(doi)
            return article_data.get('reference', [])
        except:
            return []

    @retry(stop=stop_after_attempt(Config.MAX_RETRIES),
           wait=wait_exponential(multiplier=1, min=Config.RETRY_DELAY, max=10))
    def get_citing_articles_from_openalex(self, doi: str) -> List[str]:
        """Получает цитирующие работы через OpenAlex API"""
        citing_dois = []
        try:
            work_id = doi.replace('/', '%2F')
            url = f"https://api.openalex.org/works/https://doi.org/{work_id}"

            response = requests.get(url, timeout=Config.REQUEST_TIMEOUT)
            self.performance_monitor.increment_request()

            if response.status_code == 200:
                data = response.json()
                cited_by_count = data.get('cited_by_count', 0)
                work_openalex_id = data.get('id', '')

                if cited_by_count > 0:
                    per_page = 200
                    total_pages = (cited_by_count + per_page - 1) // per_page

                    for page in range(1, total_pages + 1):
                        citing_url = f"https://api.openalex.org/works?filter=cites:{work_openalex_id}&per-page={per_page}&page={page}"

                        response = requests.get(citing_url, timeout=Config.REQUEST_TIMEOUT)
                        self.performance_monitor.increment_request()

                        if response.status_code == 200:
                            citing_data = response.json()
                            results = citing_data.get('results', [])
                            results_count = len(results)

                            for work in results:
                                if work.get('doi'):
                                    citing_dois.append(work['doi'])

                            if results_count < per_page:
                                break

                            time.sleep(Config.DELAY_BETWEEN_REQUESTS)

                        else:
                            if response.status_code == 429:
                                time.sleep(3)
                                page -= 1
                            else:
                                break

                        if len(citing_dois) >= cited_by_count:
                            break

            return citing_dois

        except Exception as e:
            return []

    def get_citing_articles_from_crossref(self, doi: str) -> List[str]:
        """Получает цитирующие работы через Crossref API"""
        citing_dois = []
        try:
            url = f"https://api.crossref.org/works/{doi}"
            response = requests.get(url, timeout=Config.REQUEST_TIMEOUT)
            self.performance_monitor.increment_request()

            if response.status_code == 200:
                data = response.json()
                if 'message' in data and 'is-referenced-by' in data['message']:
                    references = data['message']['is-referenced-by']
                    for ref in references:
                        if isinstance(ref, dict) and 'DOI' in ref:
                            citing_dois.append(ref['DOI'])

        except Exception as e:
            pass

        return citing_dois

    def find_citing_articles(self, doi_list: List[str]) -> Dict[str, Dict]:
        """Основная функция для поиска цитирующих статей"""
        results = {}

        for i, doi in enumerate(doi_list, 1):
            doi = doi.strip()
            if not doi:
                continue

            openalex_citations = self.get_citing_articles_from_openalex(doi)
            crossref_citations = self.get_citing_articles_from_crossref(doi)

            all_citations = list(set(openalex_citations + crossref_citations))

            results[doi] = {
                'count': len(all_citations),
                'citing_dois': all_citations
            }

            time.sleep(Config.DELAY_BETWEEN_REQUESTS)

        return results

    def process_citing_articles_sequential(self, doi_list: List[str]) -> tuple[pd.DataFrame, pd.DataFrame, Dict[str, Dict], List[str]]:
        """Обрабатывает цитирующие статьи последовательно с полной статистикой"""
        self.performance_monitor.start()

        citing_results = self.find_citing_articles(doi_list)

        all_citing_articles_data = []
        citing_articles_details = []
        all_citing_titles = []

        # ИЗМЕНЕНИЕ: Собираем все связи с сохранением source_doi
        all_citing_connections = []
        for source_doi, source_data in citing_results.items():
            for citing_doi in source_data['citing_dois']:
                all_citing_connections.append({
                    'source_doi': source_doi,
                    'citing_doi': citing_doi
                })

        # ИЗМЕНЕНИЕ: Кэшируем данные уникальных цитирующих статей
        all_citing_dois = set(conn['citing_doi'] for conn in all_citing_connections)
        citing_data_cache = {}

        for citing_doi in tqdm(all_citing_dois, desc="Обработка цитирующих статей"):
            try:
                article_data = self.get_combined_article_data(citing_doi)
                citing_data_cache[citing_doi] = article_data
                all_citing_titles.append(article_data['title'])
                time.sleep(Config.DELAY_BETWEEN_REQUESTS)
            except Exception as e:
                citing_data_cache[citing_doi] = {
                    'title': 'Error', 'authors': 'Error', 'authors_with_initials': 'Error',
                    'author_count': 0, 'year': 'Unknown', 'journal_full_name': 'Error',
                    'journal_abbreviation': 'Error', 'publisher': 'Error',
                    'citation_count_crossref': 0, 'citation_count_openalex': 0,
                    'years_since_publication': 1, 'affiliations': 'Error', 'countries': 'Error',
                    'altmetric_score': 0, 'number_of_mentions': 0, 'x_mentions': 0,
                    'rss_blogs': 0, 'unique_accounts': 0
                }
                all_citing_titles.append('Error')

        # ИЗМЕНЕНИЕ: Создаем citing_articles_df со ВСЕМИ связями
        for connection in all_citing_connections:
            citing_doi = connection['citing_doi']
            source_doi = connection['source_doi']

            article_data = citing_data_cache.get(citing_doi, {})
            citing_row = {
                'source_doi': source_doi,  # ПРАВИЛЬНО - сохраняем источник
                'position': 'N/A',
                'doi': citing_doi,
                'title': article_data.get('title', 'Unknown'),
                'authors': article_data.get('authors', 'Unknown'),
                'authors_with_initials': article_data.get('authors_with_initials', 'Unknown'),
                'author_count': article_data.get('author_count', 0),
                'year': article_data.get('year', 'Unknown'),
                'journal_full_name': article_data.get('journal_full_name', 'Unknown'),
                'journal_abbreviation': article_data.get('journal_abbreviation', 'Unknown'),
                'publisher': article_data.get('publisher', 'Unknown'),
                'citation_count_crossref': article_data.get('citation_count_crossref', 0),
                'citation_count_openalex': article_data.get('citation_count_openalex', 0),
                'annual_citation_rate_crossref': self.safe_calculate_annual_citation_rate(
                    article_data.get('citation_count_crossref', 0), article_data.get('publication_year')
                ),
                'annual_citation_rate_openalex': self.safe_calculate_annual_citation_rate(
                    article_data.get('citation_count_openalex', 0), article_data.get('publication_year')
                ),
                'years_since_publication': article_data.get('years_since_publication', 1),
                'affiliations': article_data.get('affiliations', 'Unknown'),
                'countries': article_data.get('countries', 'Unknown'),
                'altmetric_score': article_data.get('altmetric_score', 0),
                'number_of_mentions': article_data.get('number_of_mentions', 0),
                'x_mentions': article_data.get('x_mentions', 0),
                'rss_blogs': article_data.get('rss_blogs', 0),
                'unique_accounts': article_data.get('unique_accounts', 0),
                'error': None
            }
            all_citing_articles_data.append(citing_row)

        # Создаем DataFrame со всеми связями
        citing_articles_df = pd.DataFrame(all_citing_articles_data) if all_citing_articles_data else pd.DataFrame()

        # Создаем details таблицу
        for source_doi, data in citing_results.items():
            for citing_doi in data['citing_dois']:
                article_data = citing_data_cache.get(citing_doi, {})
                citing_articles_details.append({
                    'source_doi': source_doi,
                    'citing_doi': citing_doi,
                    'citing_title': article_data.get('title', 'Unknown'),
                    'citing_authors': article_data.get('authors_with_initials', 'Unknown'),
                    'citing_year': article_data.get('year', 'Unknown'),
                    'citing_journal': article_data.get('journal_abbreviation', 'Unknown'),
                    'citation_count': article_data.get('citation_count_openalex', 0),
                    'altmetric_score': article_data.get('altmetric_score', 0),
                    'number_of_mentions': article_data.get('number_of_mentions', 0)
                })

        citing_details_df = pd.DataFrame(citing_articles_details) if citing_articles_details else pd.DataFrame()

        return citing_articles_df, citing_details_df, citing_results, all_citing_titles

    def get_unique_citations(self, citations_df: pd.DataFrame) -> pd.DataFrame:
        """Получает уникальные цитирующие статьи"""
        if citations_df.empty:
            return pd.DataFrame()

        cache_key = id(citations_df)
        if cache_key not in self._unique_citations_cache:
            citations_df['citation_id'] = citations_df['doi'].fillna('') + '|' + citations_df['title'].fillna('')
            unique_df = citations_df.drop_duplicates(subset=['citation_id'], keep='first').drop(columns=['citation_id'])
            self._unique_citations_cache[cache_key] = unique_df
        return self._unique_citations_cache[cache_key]

    def find_duplicate_citations(self, citations_df: pd.DataFrame) -> pd.DataFrame:
        """Находит дублирующиеся цитирующие статьи (статьи, которые цитируют несколько анализируемых работ)"""
        try:
            if citations_df.empty:
                return pd.DataFrame()

            citations_df['citation_id'] = citations_df['doi'].fillna('') + '|' + citations_df['title'].fillna('')

            # Считаем, сколько разных исходных статей цитирует каждая цитирующая статья
            citation_counts = citations_df.groupby('citation_id')['source_doi'].nunique().reset_index()
            duplicate_citation_ids = citation_counts[citation_counts['source_doi'] > 1]['citation_id']

            if duplicate_citation_ids.empty:
                columns = list(citations_df.columns) + ['frequency']
                columns.remove('citation_id')
                return pd.DataFrame(columns=columns)

            # Создаем маппинг citation_id -> frequency (количество цитируемых исходных статей)
            frequency_map = citation_counts.set_index('citation_id')['source_doi'].to_dict()

            # Отбираем дублирующиеся цитаты и оставляем по одной записи на каждую цитирующую статью
            duplicates = citations_df[citations_df['citation_id'].isin(duplicate_citation_ids)].copy()
            duplicates = duplicates.drop_duplicates(subset=['citation_id'], keep='first')

            # Фильтруем некорректные записи
            duplicates = duplicates[~((duplicates['doi'].isna()) & (duplicates['title'] == 'Unknown'))]

            # Добавляем колонку с частотой
            duplicates['frequency'] = duplicates['citation_id'].map(frequency_map)
            duplicates = duplicates.drop(columns=['citation_id'])

            return duplicates.sort_values(['frequency', 'doi'], ascending=[False, True])
        except Exception as e:
            st.error(f"Error in find_duplicate_citations: {e}")
            return pd.DataFrame()

    def analyze_citation_authors_frequency(self, citations_df: pd.DataFrame) -> pd.DataFrame:
        """Анализ частоты авторов в цитирующих статьях"""
        try:
            if citations_df.empty:
                return pd.DataFrame()

            total_citations = len(citations_df)
            unique_df = self.get_unique_citations(citations_df)
            total_unique = len(unique_df)

            authors_total = citations_df['authors_with_initials'].str.split(',', expand=True).stack()
            authors_total = authors_total[authors_total.str.strip().isin(['Unknown', 'Error']) == False]
            author_freq_total = authors_total.value_counts().reset_index()
            author_freq_total.columns = ['author_with_initial', 'frequency_total']
            author_freq_total['percentage_total'] = round(author_freq_total['frequency_total'] / total_citations * 100, 2)

            authors_unique = unique_df['authors_with_initials'].str.split(',', expand=True).stack()
            authors_unique = authors_unique[authors_unique.str.strip().isin(['Unknown', 'Error']) == False]
            author_freq_unique = authors_unique.value_counts().reset_index()
            author_freq_unique.columns = ['author_with_initial', 'frequency_unique']
            author_freq_unique['percentage_unique'] = round(author_freq_unique['frequency_unique'] / total_unique * 100, 2)

            author_freq = author_freq_total.merge(author_freq_unique, on='author_with_initial', how='outer').fillna(0)
            return author_freq[['author_with_initial', 'frequency_total', 'percentage_total', 'frequency_unique', 'percentage_unique']].sort_values('frequency_total', ascending=False)
        except Exception as e:
            return pd.DataFrame()

    def analyze_citation_journals_frequency(self, citations_df: pd.DataFrame) -> pd.DataFrame:
        """Анализ частоты журналов в цитирующих статьях с дополнительными метриками"""
        try:
            if citations_df.empty:
                return pd.DataFrame()

            total_citations = len(citations_df)
            unique_df = self.get_unique_citations(citations_df)
            total_unique = len(unique_df)

            journals_total = citations_df['journal_abbreviation']
            journals_total = journals_total[journals_total.isin(['Unknown', 'Error']) == False]
            journal_freq_total = journals_total.value_counts().reset_index()
            journal_freq_total.columns = ['journal_abbreviation', 'frequency_total']
            journal_freq_total['percentage_total'] = round(journal_freq_total['frequency_total'] / total_citations * 100, 2)

            journals_unique = unique_df['journal_abbreviation']
            journals_unique = journals_unique[journals_unique.isin(['Unknown', 'Error']) == False]
            journal_freq_unique = journals_unique.value_counts().reset_index()
            journal_freq_unique.columns = ['journal_abbreviation', 'frequency_unique']
            journal_freq_unique['percentage_unique'] = round(journal_freq_unique['frequency_unique'] / total_unique * 100, 2)

            journal_freq = journal_freq_total.merge(journal_freq_unique, on='journal_abbreviation', how='outer').fillna(0)

            # Добавляем дополнительные метрики для цитирований
            journal_citation_metrics = []
            for journal in journal_freq['journal_abbreviation']:
                journal_articles = unique_df[unique_df['journal_abbreviation'] == journal]

                total_crossref_citations = journal_articles['citation_count_crossref'].sum()
                total_openalex_citations = journal_articles['citation_count_openalex'].sum()

                avg_crossref_citations = journal_articles['citation_count_crossref'].mean() if len(journal_articles) > 0 else 0
                avg_openalex_citations = journal_articles['citation_count_openalex'].mean() if len(journal_articles) > 0 else 0

                journal_citation_metrics.append({
                    'journal_abbreviation': journal,
                    'total_crossref_citations': total_crossref_citations,
                    'total_openalex_citations': total_openalex_citations,
                    'avg_crossref_citations': round(avg_crossref_citations, 2),
                    'avg_openalex_citations': round(avg_openalex_citations, 2)
                })

            journal_metrics_df = pd.DataFrame(journal_citation_metrics)
            journal_freq = journal_freq.merge(journal_metrics_df, on='journal_abbreviation', how='left').fillna(0)

            return journal_freq[['journal_abbreviation', 'frequency_total', 'percentage_total',
                               'frequency_unique', 'percentage_unique', 'total_crossref_citations',
                               'total_openalex_citations', 'avg_crossref_citations', 'avg_openalex_citations']].sort_values('frequency_total', ascending=False)
        except Exception as e:
            return pd.DataFrame()

    def analyze_citation_affiliations_frequency(self, citations_df: pd.DataFrame) -> pd.DataFrame:
        """Анализ частоты аффилиаций в цитирующих статьях с группировкой"""
        try:
            if citations_df.empty:
                return pd.DataFrame()

            total_citations = len(citations_df)
            unique_df = self.get_unique_citations(citations_df)
            total_unique = len(unique_df)

            all_affiliations = []
            for affil_string in citations_df['affiliations']:
                if pd.isna(affil_string) or affil_string in ['Unknown', 'Error', '']:
                    continue

                try:
                    affil_list = affil_string.split(';')
                    for affil in affil_list:
                        clean_affil = affil.strip()
                        if clean_affil and clean_affil not in ['Unknown', 'Error']:
                            all_affiliations.append(clean_affil)
                except Exception:
                    pass

            if not all_affiliations:
                return pd.DataFrame(columns=['affiliation', 'frequency_total', 'percentage_total', 'frequency_unique', 'percentage_unique'])

            affiliation_frequencies, grouped_organizations = self.fast_affiliation_processor.process_affiliations_list_fast(all_affiliations)

            affil_data = []
            for affil, freq in affiliation_frequencies.items():
                percentage_total = round(freq / total_citations * 100, 2) if total_citations > 0 else 0
                affil_data.append({
                    'affiliation': affil,
                    'frequency_total': freq,
                    'percentage_total': percentage_total
                })

            unique_affiliations = []
            for affil_string in unique_df['affiliations']:
                if affil_string and affil_string not in ['Unknown', 'Error']:
                    affil_list = affil_string.split(';')
                    unique_affiliations.extend([affil.strip() for affil in affil_list if affil.strip()])

            if unique_affiliations:
                unique_frequencies, _ = self.fast_affiliation_processor.process_affiliations_list_fast(unique_affiliations)

                affil_df = pd.DataFrame(affil_data)
                for affil, freq in unique_frequencies.items():
                    percentage_unique = round(freq / total_unique * 100, 2) if total_unique > 0 else 0
                    if affil in affil_df['affiliation'].values:
                        affil_df.loc[affil_df['affiliation'] == affil, 'frequency_unique'] = freq
                        affil_df.loc[affil_df['affiliation'] == affil, 'percentage_unique'] = percentage_unique
                    else:
                        affil_df = pd.concat([affil_df, pd.DataFrame([{
                            'affiliation': affil,
                            'frequency_total': 0,
                            'percentage_total': 0,
                            'frequency_unique': freq,
                            'percentage_unique': percentage_unique
                        }])], ignore_index=True)
            else:
                affil_df = pd.DataFrame(affil_data)
                affil_df['frequency_unique'] = affil_df['frequency_total']
                affil_df['percentage_unique'] = affil_df['percentage_total']

            affil_df = affil_df.fillna(0)
            return affil_df[['affiliation', 'frequency_total', 'percentage_total', 'frequency_unique', 'percentage_unique']].sort_values('frequency_total', ascending=False)

        except Exception as e:
            return pd.DataFrame()

    def analyze_citation_countries_frequency(self, citations_df: pd.DataFrame) -> pd.DataFrame:
        """Анализ частоты стран в цитирующих статьях с разделением на отдельные страны и коллаборации"""
        try:
            if citations_df.empty:
                return pd.DataFrame()

            total_citations = len(citations_df)
            unique_df = self.get_unique_citations(citations_df)
            total_unique = len(unique_df)

            # Анализ отдельных стран
            single_countries = []
            collaborations = []

            for countries in citations_df['countries']:
                if countries not in ['Unknown', 'Error']:
                    country_list = [c.strip() for c in countries.split(';')]
                    if len(country_list) == 1:
                        single_countries.extend(country_list)
                    else:
                        collaborations.append(countries)

            # Частотность отдельных стран
            single_country_counter = Counter(single_countries)
            single_country_freq = pd.DataFrame({
                'countries': list(single_country_counter.keys()),
                'type': ['single'] * len(single_country_counter),
                'frequency_total': list(single_country_counter.values())
            })
            single_country_freq['percentage_total'] = round(single_country_freq['frequency_total'] / total_citations * 100, 2)

            # Частотность коллабораций
            collaboration_counter = Counter(collaborations)
            collaboration_freq = pd.DataFrame({
                'countries': list(collaboration_counter.keys()),
                'type': ['collaboration'] * len(collaboration_counter),
                'frequency_total': list(collaboration_counter.values())
            })
            collaboration_freq['percentage_total'] = round(collaboration_freq['frequency_total'] / total_citations * 100, 2)

            # Объединяем
            country_freq_total = pd.concat([single_country_freq, collaboration_freq], ignore_index=True)

            # Аналогично для уникальных статей
            single_countries_unique = []
            collaborations_unique = []

            for countries in unique_df['countries']:
                if countries not in ['Unknown', 'Error']:
                    country_list = [c.strip() for c in countries.split(';')]
                    if len(country_list) == 1:
                        single_countries_unique.extend(country_list)
                    else:
                        collaborations_unique.append(countries)

            single_country_counter_unique = Counter(single_countries_unique)
            collaboration_counter_unique = Counter(collaborations_unique)

            single_country_freq_unique = pd.DataFrame({
                'countries': list(single_country_counter_unique.keys()),
                'frequency_unique': list(single_country_counter_unique.values())
            })
            single_country_freq_unique['percentage_unique'] = round(single_country_freq_unique['frequency_unique'] / total_unique * 100, 2)

            collaboration_freq_unique = pd.DataFrame({
                'countries': list(collaboration_counter_unique.keys()),
                'frequency_unique': list(collaboration_counter_unique.values())
            })
            collaboration_freq_unique['percentage_unique'] = round(collaboration_freq_unique['frequency_unique'] / total_unique * 100, 2)

            country_freq_unique = pd.concat([
                single_country_freq_unique.assign(type='single'),
                collaboration_freq_unique.assign(type='collaboration')
            ], ignore_index=True)

            country_freq = country_freq_total.merge(country_freq_unique, on=['countries', 'type'], how='outer').fillna(0)

            return country_freq[['countries', 'type', 'frequency_total', 'percentage_total', 'frequency_unique', 'percentage_unique']].sort_values('frequency_total', ascending=False)
        except Exception as e:
            return pd.DataFrame()

    def analyze_citation_year_distribution(self, citations_df: pd.DataFrame) -> pd.DataFrame:
        """Анализ распределения по годам для цитирующих статей (от новых к старым)"""
        try:
            if citations_df.empty:
                return pd.DataFrame()

            total_citations = len(citations_df)
            unique_df = self.get_unique_citations(citations_df)
            total_unique = len(unique_df)

            years_total = pd.to_numeric(citations_df['year'], errors='coerce')
            years_total = years_total[years_total.notna() & years_total.between(1900, 2026)].astype(int)
            year_counts_total = years_total.value_counts().reset_index()
            year_counts_total.columns = ['year', 'frequency_total']
            year_counts_total['percentage_total'] = round(year_counts_total['frequency_total'] / total_citations * 100, 2)

            years_unique = pd.to_numeric(unique_df['year'], errors='coerce')
            years_unique = years_unique[years_unique.notna() & years_unique.between(1900, 2026)].astype(int)
            year_counts_unique = years_unique.value_counts().reset_index()
            year_counts_unique.columns = ['year', 'frequency_unique']
            year_counts_unique['percentage_unique'] = round(year_counts_unique['frequency_unique'] / total_unique * 100, 2)

            year_counts = year_counts_total.merge(year_counts_unique, on='year', how='outer').fillna(0)
            return year_counts[['year', 'frequency_total', 'percentage_total', 'frequency_unique', 'percentage_unique']].sort_values('year', ascending=False)
        except Exception as e:
            return pd.DataFrame()

    def analyze_citation_five_year_periods(self, citations_df: pd.DataFrame) -> pd.DataFrame:
        """Анализ пятилетних периодов для цитирующих статей (от новых к старым)"""
        try:
            if citations_df.empty:
                return pd.DataFrame()

            total_citations = len(citations_df)
            unique_df = self.get_unique_citations(citations_df)
            total_unique = len(unique_df)

            start_year = 1900
            current_year = datetime.now().year + 4
            period_starts = list(range(start_year, current_year + 1, 5))
            bins = period_starts + [period_starts[-1] + 5]
            labels = [f"{s}-{s+4}" for s in period_starts]

            years_total = pd.to_numeric(citations_df['year'], errors='coerce')
            years_total = years_total[years_total.notna() & years_total.between(1900, current_year)].astype(int)
            period_counts_total = pd.cut(years_total, bins=bins, labels=labels, right=False).astype(str)
            period_df_total = period_counts_total.value_counts().reset_index()
            period_df_total.columns = ['period', 'frequency_total']
            period_df_total['percentage_total'] = round(period_df_total['frequency_total'] / total_citations * 100, 2)
            period_df_total['period'] = period_df_total['period'].astype(str)

            years_unique = pd.to_numeric(unique_df['year'], errors='coerce')
            years_unique = years_unique[years_unique.notna() & years_unique.between(1900, current_year)].astype(int)
            period_counts_unique = pd.cut(years_unique, bins=bins, labels=labels, right=False).astype(str)
            period_df_unique = period_counts_unique.value_counts().reset_index()
            period_df_unique.columns = ['period', 'frequency_unique']
            period_df_unique['percentage_unique'] = round(period_df_unique['frequency_unique'] / total_unique * 100, 2)
            period_df_unique['period'] = period_df_unique['period'].astype(str)

            period_df = period_df_total.merge(period_df_unique, on='period', how='outer').fillna(0)
            return period_df[['period', 'frequency_total', 'percentage_total', 'frequency_unique', 'percentage_unique']].sort_values('period', ascending=False)
        except Exception as e:
            return pd.DataFrame()

    def save_citation_analysis_to_excel(self, citing_articles_df: pd.DataFrame, citing_details_df: pd.DataFrame,
                                      doi_list: List[str], citing_results: Dict, all_citing_titles: List[str]) -> BytesIO:
        """Сохраняет полный анализ цитирующих статей в Excel"""
        try:
            timestamp = int(time.time())
            excel_buffer = BytesIO()

            wb = Workbook()

            # Сначала создаем вкладку Report_Summary
            ws_summary = wb.active
            ws_summary.title = 'Report_Summary_Citations'

            wb.remove(wb.active)

            try:
                unique_citations_df = self.get_unique_citations(citing_articles_df)
            except Exception as e:
                unique_citations_df = pd.DataFrame()

            try:
                duplicate_citations_df = self.find_duplicate_citations(citing_articles_df)
            except Exception as e:
                duplicate_citations_df = pd.DataFrame()

            stats = self.performance_monitor.get_stats()

            try:
                content_freq, compound_freq, scientific_freq = self.analyze_titles(all_citing_titles)
            except Exception as e:
                content_freq, compound_freq, scientific_freq = Counter(), Counter(), Counter()

            try:
                total_citation_relationships = len(citing_articles_df) if not citing_articles_df.empty else 0
                total_unique_citations = len(unique_citations_df) if not unique_citations_df.empty else 0

                countries_with_data = citing_articles_df[citing_articles_df['countries'].isin(['Unknown', 'Error']) == False]
                countries_percentage = (len(countries_with_data) / total_citation_relationships * 100) if total_citation_relationships > 0 else 0

                affiliations_with_data = citing_articles_df[citing_articles_df['affiliations'].isin(['Unknown', 'Error']) == False]
                affiliations_percentage = (len(affiliations_with_data) / total_citation_relationships * 100) if total_citation_relationships > 0 else 0

            except Exception as e:
                total_citation_relationships = 0
                total_unique_citations = 0
                countries_percentage = 0
                affiliations_percentage = 0

            citing_info = ""
            if citing_results:
                citing_info = f"\nCitations per source article:"
                for doi, data in citing_results.items():
                    citing_info += f"\n  - {doi}: {data['count']} citations"

            summary_content = f"""@MedvDmitry production

CITATION ANALYSIS REPORT (CITING ARTICLES)

Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

ANALYSIS OVERVIEW
=================
Total source articles: {len(doi_list)}
Total citation relationships: {total_citation_relationships}
Total unique citing articles: {total_unique_citations}
Successful citations: {len(citing_articles_df[citing_articles_df['error'].isna()]) if not citing_articles_df.empty else 0}
Failed citations: {len(citing_articles_df[citing_articles_df['error'].notna()]) if not citing_articles_df.empty else 0}
Unique authors: {len(self.analyze_citation_authors_frequency(citing_articles_df)) if not citing_articles_df.empty else 0}
Unique journals: {len(self.analyze_citation_journals_frequency(citing_articles_df)) if not citing_articles_df.empty else 0}
Unique affiliations: {len(self.analyze_citation_affiliations_frequency(citing_articles_df)) if not citing_articles_df.empty else 0}
Unique countries: {len(self.analyze_citation_countries_frequency(citing_articles_df)) if not citing_articles_df.empty else 0}
Duplicate citations: {len(duplicate_citations_df) if not duplicate_citations_df.empty else 0}

DATA COMPLETENESS
=================
Articles with country data: {countries_percentage:.1f}%
Articles with affiliation data: {affiliations_percentage:.1f}%

AFFILIATION PROCESSING
======================
Affiliations normalized and grouped by organization
Similar affiliations merged together
Frequency counts reflect grouped organizations

PERFORMANCE STATISTICS
======================
Total processing time: {stats.get('elapsed_seconds', 0):.2f} seconds ({stats.get('elapsed_minutes', 0):.2f} minutes)
Total API requests: {stats.get('total_requests', 0)}
Requests per second: {stats.get('requests_per_second', 0):.2f}
{citing_info}

DATA QUALITY NOTES
==================
Analysis focuses on articles that cite the source articles
Combined data from Crossref and OpenAlex improves completeness
All standard statistical analyses performed (authors, journals, countries, etc.)
Error handling ensures report generation even with partial data
Duplicate citations show articles that cite multiple source articles
Affiliations normalized and grouped for consistent organization names
Altmetric metrics included for social media and online attention analysis
"""

            # Создаем вкладку Report_Summary первой
            ws_summary = wb.create_sheet('Report_Summary_Citations')
            for line in summary_content.split('\n'):
                ws_summary.append([line])

            # Удаляем колонку authors_surnames и добавляем author_count для основных таблиц
            if not citing_articles_df.empty and 'authors_surnames' in citing_articles_df.columns:
                citing_articles_df = citing_articles_df.drop(columns=['authors_surnames'])
            if not unique_citations_df.empty and 'authors_surnames' in unique_citations_df.columns:
                unique_citations_df = unique_citations_df.drop(columns=['authors_surnames'])
            if not duplicate_citations_df.empty and 'authors_surnames' in duplicate_citations_df.columns:
                duplicate_citations_df = duplicate_citations_df.drop(columns=['authors_surnames'])

            sheets_data = [
                ('Source_Articles_Citations', citing_details_df),
                ('All_Citations', citing_articles_df),
                ('All_Unique_Citations', unique_citations_df),
                ('Duplicate_Citations', duplicate_citations_df)
            ]

            analysis_methods = [
                ('Author_Frequency_Citations', self.analyze_citation_authors_frequency),
                ('Journal_Frequency_Citations', self.analyze_citation_journals_frequency),
                ('Affiliation_Frequency_Citations', self.analyze_citation_affiliations_frequency),
                ('Country_Frequency_Citations', self.analyze_citation_countries_frequency),
                ('Year_Distribution_Citations', self.analyze_citation_year_distribution),
                ('5_Years_Period_Citations', self.analyze_citation_five_year_periods)
            ]

            for sheet_name, method in analysis_methods:
                try:
                    result_df = method(citing_articles_df)
                    sheets_data.append((sheet_name, result_df))
                except Exception as e:
                    sheets_data.append((sheet_name, pd.DataFrame()))

            try:
                title_word_data = []
                for i, (word, count) in enumerate(content_freq.most_common(50), 1):
                    title_word_data.append({'Category': 'Content_Words', 'Rank': i, 'Word': word, 'Frequency': count})
                for i, (word, count) in enumerate(compound_freq.most_common(50), 1):
                    title_word_data.append({'Category': 'Compound_Words', 'Rank': i, 'Word': word, 'Frequency': count})
                for i, (word, count) in enumerate(scientific_freq.most_common(50), 1):
                    title_word_data.append({'Category': 'Scientific_Stopwords', 'Rank': i, 'Word': word, 'Frequency': count})

                title_word_df = pd.DataFrame(title_word_data)
                sheets_data.append(('Title_Word_Frequency_Citations', title_word_df))
            except Exception as e:
                sheets_data.append(('Title_Word_Frequency_Citations', pd.DataFrame()))

            for sheet_name, df in sheets_data:
                try:
                    if not df.empty:
                        ws = wb.create_sheet(sheet_name)
                        for r in dataframe_to_rows(df, index=False, header=True):
                            ws.append(r)
                    else:
                        ws = wb.create_sheet(sheet_name)
                        ws.append([f"No data available for {sheet_name}"])
                except Exception as e:
                    pass

            wb.save(excel_buffer)
            excel_buffer.seek(0)
            return excel_buffer

        except Exception as e:
            try:
                timestamp = int(time.time())
                excel_buffer = BytesIO()
                wb = Workbook()
                ws = wb.active
                ws.title = "Error_Report_Citations"
                ws.append(["ERROR REPORT - CITATION ANALYSIS"])
                ws.append([f"Critical error during citation analysis: {str(e)}"])
                ws.append([f"Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
                ws.append(["DOIs processed:", ', '.join(doi_list)])
                wb.save(excel_buffer)
                excel_buffer.seek(0)
                return excel_buffer
            except:
                return BytesIO()

    def process_doi_sequential(self, doi_list: List[str]) -> tuple[pd.DataFrame, pd.DataFrame, int, int, List[str]]:
        """Process DOIs sequentially to avoid API overload"""
        self.performance_monitor.start()

        all_references = []
        for doi in tqdm(doi_list, desc="Collecting references"):
            try:
                references = self.get_references_from_crossref(doi)
                for i, ref in enumerate(references):
                    all_references.append({
                        'source_doi': doi,
                        'position': i + 1,
                        'ref': ref
                    })
                time.sleep(Config.DELAY_BETWEEN_REQUESTS)
            except Exception as e:
                pass

        unique_dois = set()
        titles_to_search = set()
        all_titles = []

        for ref_data in all_references:
            ref = ref_data['ref']
            ref_doi = ref.get('DOI')
            title = ref.get('article-title', 'Unknown')
            all_titles.append(title)

            if ref_doi and self.validate_doi(ref_doi):
                unique_dois.add(ref_doi)
            elif title != 'Unknown':
                titles_to_search.add(title)

        title_to_doi = {}
        for title in tqdm(list(titles_to_search), desc="Searching DOIs by title"):
            doi = self.quick_doi_search(title)
            if doi and self.validate_doi(doi):
                normalized_doi = self.normalize_doi(doi)
                title_to_doi[title] = normalized_doi
                unique_dois.add(normalized_doi)
            time.sleep(Config.DELAY_BETWEEN_REQUESTS)

        for doi in tqdm(list(unique_dois), desc="Processing unique DOIs"):
            if doi not in self.unique_ref_data_cache:
                try:
                    article_data = self.get_combined_article_data(doi)
                    self.unique_ref_data_cache[doi] = {
                        'doi': doi,
                        'title': article_data['title'],
                        'authors': article_data['authors'],
                        'authors_surnames': article_data['authors_surnames'],
                        'authors_with_initials': article_data['authors_with_initials'],
                        'author_count': article_data['author_count'],
                        'year': article_data['year'],
                        'journal_full_name': article_data['journal_full_name'],
                        'journal_abbreviation': article_data['journal_abbreviation'],
                        'publisher': article_data['publisher'],
                        'citation_count_crossref': article_data['citation_count_crossref'],
                        'citation_count_openalex': article_data['citation_count_openalex'],
                        'affiliations': article_data['affiliations'],
                        'countries': article_data['countries'],
                        'publication_year': article_data.get('publication_year'),
                        'years_since_publication': article_data['years_since_publication'],
                        'altmetric_score': article_data['altmetric_score'],
                        'number_of_mentions': article_data['number_of_mentions'],
                        'x_mentions': article_data['x_mentions'],
                        'rss_blogs': article_data['rss_blogs'],
                        'unique_accounts': article_data['unique_accounts']
                    }
                except Exception as e:
                    self.unique_ref_data_cache[doi] = {
                        'doi': doi, 'title': 'Unknown', 'authors': 'Error',
                        'authors_surnames': 'Error', 'authors_with_initials': 'Error', 'author_count': 0,
                        'year': 'Unknown', 'journal_full_name': 'Error',
                        'journal_abbreviation': 'Error', 'publisher': 'Error',
                        'citation_count_crossref': 'N/A', 'citation_count_openalex': 'N/A',
                        'affiliations': 'Error', 'countries': 'Error', 'error': str(e),
                        'altmetric_score': 0, 'number_of_mentions': 0, 'x_mentions': 0,
                        'rss_blogs': 0, 'unique_accounts': 0
                    }
                time.sleep(Config.DELAY_BETWEEN_REQUESTS)

        results = []
        source_articles = []

        for doi in tqdm(doi_list, desc="Processing source articles"):
            try:
                source_data = self.get_combined_article_data(doi)
                source_row = {
                    'source_doi': doi,
                    'position': None,
                    'doi': doi,
                    'title': source_data['title'],
                    'authors': source_data['authors'],
                    'authors_with_initials': source_data['authors_with_initials'],
                    'author_count': source_data['author_count'],
                    'year': source_data['year'],
                    'journal_full_name': source_data['journal_full_name'],
                    'journal_abbreviation': source_data['journal_abbreviation'],
                    'publisher': source_data['publisher'],
                    'citation_count_crossref': source_data['citation_count_crossref'],
                    'citation_count_openalex': source_data['citation_count_openalex'],
                    'annual_citation_rate_crossref': self.safe_calculate_annual_citation_rate(
                        source_data['citation_count_crossref'], source_data.get('publication_year')
                    ),
                    'annual_citation_rate_openalex': self.safe_calculate_annual_citation_rate(
                        source_data['citation_count_openalex'], source_data.get('publication_year')
                    ),
                    'years_since_publication': source_data['years_since_publication'],
                    'affiliations': source_data['affiliations'],
                    'countries': source_data['countries'],
                    'altmetric_score': source_data['altmetric_score'],
                    'number_of_mentions': source_data['number_of_mentions'],
                    'x_mentions': source_data['x_mentions'],
                    'rss_blogs': source_data['rss_blogs'],
                    'unique_accounts': source_data['unique_accounts'],
                    'error': None
                }
                source_articles.append(source_row)
            except Exception as e:
                source_articles.append({
                    'source_doi': doi, 'position': None, 'doi': doi, 'title': 'Unknown',
                    'authors': 'Error', 'authors_with_initials': 'Error', 'author_count': 0,
                    'year': 'Unknown', 'journal_full_name': 'Error', 'journal_abbreviation': 'Error',
                    'publisher': 'Error', 'citation_count_crossref': 'N/A', 'citation_count_openalex': 'N/A',
                    'annual_citation_rate_crossref': 'N/A', 'annual_citation_rate_openalex': 'N/A',
                    'years_since_publication': 'N/A', 'affiliations': 'Error', 'countries': 'Error',
                    'altmetric_score': 0, 'number_of_mentions': 0, 'x_mentions': 0, 'rss_blogs': 0, 'unique_accounts': 0,
                    'error': str(e)
                })

            article_refs = [ref for ref in all_references if ref['source_doi'] == doi]
            for ref_data in article_refs:
                ref = ref_data['ref']
                position = ref_data['position']
                ref_doi = ref.get('DOI')
                title = ref.get('article-title', 'Unknown')

                if ref_doi and self.validate_doi(ref_doi) and ref_doi in self.unique_ref_data_cache:
                    ref_info = self.unique_ref_data_cache[ref_doi].copy()
                    ref_row = {
                        'source_doi': doi,
                        'position': position,
                        'doi': ref_doi,
                        'title': ref_info['title'],
                        'authors': ref_info['authors'],
                        'authors_with_initials': ref_info['authors_with_initials'],
                        'author_count': ref_info['author_count'],
                        'year': ref_info['year'],
                        'journal_full_name': ref_info['journal_full_name'],
                        'journal_abbreviation': ref_info['journal_abbreviation'],
                        'publisher': ref_info['publisher'],
                        'citation_count_crossref': ref_info['citation_count_crossref'],
                        'citation_count_openalex': ref_info['citation_count_openalex'],
                        'annual_citation_rate_crossref': self.safe_calculate_annual_citation_rate(
                            ref_info['citation_count_crossref'], ref_info.get('publication_year')
                        ),
                        'annual_citation_rate_openalex': self.safe_calculate_annual_citation_rate(
                            ref_info['citation_count_openalex'], ref_info.get('publication_year')
                        ),
                        'years_since_publication': ref_info['years_since_publication'],
                        'affiliations': ref_info['affiliations'],
                        'countries': ref_info['countries'],
                        'altmetric_score': ref_info['altmetric_score'],
                        'number_of_mentions': ref_info['number_of_mentions'],
                        'x_mentions': ref_info['x_mentions'],
                        'rss_blogs': ref_info['rss_blogs'],
                        'unique_accounts': ref_info['unique_accounts'],
                        'error': None
                    }
                    results.append(ref_row)
                else:
                    found_doi = title_to_doi.get(title)
                    if found_doi and found_doi in self.unique_ref_data_cache:
                        ref_info = self.unique_ref_data_cache[found_doi].copy()
                        ref_row = {
                            'source_doi': doi,
                            'position': position,
                            'doi': found_doi,
                            'title': ref_info['title'],
                            'authors': ref_info['authors'],
                            'authors_with_initials': ref_info['authors_with_initials'],
                            'author_count': ref_info['author_count'],
                            'year': ref_info['year'],
                            'journal_full_name': ref_info['journal_full_name'],
                            'journal_abbreviation': ref_info['journal_abbreviation'],
                            'publisher': ref_info['publisher'],
                            'citation_count_crossref': ref_info['citation_count_crossref'],
                            'citation_count_openalex': ref_info['citation_count_openalex'],
                            'annual_citation_rate_crossref': self.safe_calculate_annual_citation_rate(
                                ref_info['citation_count_crossref'], ref_info.get('publication_year')
                            ),
                            'annual_citation_rate_openalex': self.safe_calculate_annual_citation_rate(
                                ref_info['citation_count_openalex'], ref_info.get('publication_year')
                            ),
                            'years_since_publication': ref_info['years_since_publication'],
                            'affiliations': ref_info['affiliations'],
                            'countries': ref_info['countries'],
                            'altmetric_score': ref_info['altmetric_score'],
                            'number_of_mentions': ref_info['number_of_mentions'],
                            'x_mentions': ref_info['x_mentions'],
                            'rss_blogs': ref_info['rss_blogs'],
                            'unique_accounts': ref_info['unique_accounts'],
                            'error': None
                        }
                        results.append(ref_row)
                    else:
                        results.append({
                            'source_doi': doi, 'position': position, 'doi': ref_doi, 'title': title,
                            'authors': 'Unknown', 'authors_with_initials': 'Unknown', 'author_count': 0,
                            'year': ref.get('year', 'Unknown'), 'journal_full_name': 'Unknown',
                            'journal_abbreviation': 'Unknown', 'publisher': 'Unknown',
                            'citation_count_crossref': 'N/A', 'citation_count_openalex': 'N/A',
                            'annual_citation_rate_crossref': 'N/A', 'annual_citation_rate_openalex': 'N/A',
                            'years_since_publication': 'N/A', 'affiliations': 'Unknown', 'countries': 'Unknown',
                            'altmetric_score': 0, 'number_of_mentions': 0, 'x_mentions': 0, 'rss_blogs': 0, 'unique_accounts': 0,
                            'error': f"Invalid or missing DOI: {ref_doi}, no match found for title '{title}'"
                        })

            time.sleep(Config.DELAY_BETWEEN_REQUESTS)

        try:
            combined_references_df = pd.DataFrame(results)
        except Exception as e:
            combined_references_df = pd.DataFrame()

        try:
            source_articles_df = pd.DataFrame(source_articles)
        except Exception as e:
            source_articles_df = pd.DataFrame()

        try:
            combined_references_df = self.enhance_incomplete_data(combined_references_df)
        except Exception as e:
            pass

        try:
            source_articles_df = self.enhance_incomplete_data(source_articles_df)
        except Exception as e:
            pass

        return combined_references_df, source_articles_df, len(all_references), len(unique_dois), all_titles

    def enhance_incomplete_data(self, references_df: pd.DataFrame) -> pd.DataFrame:
        """Enhance references with incomplete data"""
        if references_df.empty:
            return references_df

        enhanced_rows = []
        incomplete_count = 0

        for index, row in tqdm(references_df.iterrows(), total=len(references_df), desc="Enhancing data"):
            doi = row['doi']

            needs_enhancement = (
                pd.isna(doi) or
                row['title'] == 'Unknown' or
                row['authors'] == 'Unknown' or
                row['affiliations'] == 'Unknown' or
                row['countries'] == 'Unknown' or
                pd.notna(row.get('error'))
            )

            if needs_enhancement and doi and self.validate_doi(doi):
                incomplete_count += 1
                try:
                    enhanced_data = self.get_combined_article_data(doi)
                    enhanced_row = {
                        'source_doi': row['source_doi'],
                        'position': row['position'],
                        'doi': doi,
                        'title': enhanced_data['title'],
                        'authors': enhanced_data['authors'],
                        'authors_with_initials': enhanced_data['authors_with_initials'],
                        'author_count': enhanced_data['author_count'],
                        'year': enhanced_data['year'],
                        'journal_full_name': enhanced_data['journal_full_name'],
                        'journal_abbreviation': enhanced_data['journal_abbreviation'],
                        'publisher': enhanced_data['publisher'],
                        'citation_count_crossref': enhanced_data['citation_count_crossref'],
                        'citation_count_openalex': enhanced_data['citation_count_openalex'],
                        'annual_citation_rate_crossref': self.safe_calculate_annual_citation_rate(
                            enhanced_data['citation_count_crossref'], enhanced_data.get('publication_year')
                        ),
                        'annual_citation_rate_openalex': self.safe_calculate_annual_citation_rate(
                            enhanced_data['citation_count_openalex'], enhanced_data.get('publication_year')
                        ),
                        'years_since_publication': enhanced_data['years_since_publication'],
                        'affiliations': enhanced_data['affiliations'],
                        'countries': enhanced_data['countries'],
                        'altmetric_score': enhanced_data['altmetric_score'],
                        'number_of_mentions': enhanced_data['number_of_mentions'],
                        'x_mentions': enhanced_data['x_mentions'],
                        'rss_blogs': enhanced_data['rss_blogs'],
                        'unique_accounts': enhanced_data['unique_accounts'],
                        'error': None
                    }
                    enhanced_rows.append(enhanced_row)
                    time.sleep(Config.DELAY_BETWEEN_REQUESTS)
                    continue
                except Exception as e:
                    pass

            enhanced_rows.append(row.to_dict())

        return pd.DataFrame(enhanced_rows)

    def reprocess_failed_references(self, failed_references_df: pd.DataFrame) -> pd.DataFrame:
        """Reprocess failed references to find missing DOIs and data"""
        if failed_references_df.empty:
            return failed_references_df

        updated_rows = []

        for index, row in tqdm(failed_references_df.iterrows(), total=len(failed_references_df), desc="Reprocessing failed"):
            original_doi = row.get('reference_doi') if 'reference_doi' in row else row.get('doi')
            error_description = row.get('error_description', '') if 'error_description' in row else row.get('error', '')

            title_match = re.search(r"title '([^']+)'|title ([^,]+)", str(error_description))
            title = None
            if title_match:
                title = next((g for g in title_match.groups() if g), None)

            if not title and 'title' in row:
                title = row['title']

            found_doi = None
            if title and title != 'Unknown':
                found_doi = self.quick_doi_search(title)
                time.sleep(Config.DELAY_BETWEEN_REQUESTS)

            if found_doi and self.validate_doi(found_doi):
                try:
                    enhanced_data = self.get_combined_article_data(found_doi)
                    enhanced_data.update({
                        'source_doi': row.get('source_doi'),
                        'position': row.get('position'),
                        'error': None,
                        'annual_citation_rate_crossref': self.safe_calculate_annual_citation_rate(
                            enhanced_data['citation_count_crossref'], enhanced_data.get('publication_year')
                        ),
                        'annual_citation_rate_openalex': self.safe_calculate_annual_citation_rate(
                            enhanced_data['citation_count_openalex'], enhanced_data.get('publication_year')
                        )
                    })
                    updated_rows.append(enhanced_data)
                    continue
                except Exception as e:
                    pass

            updated_row = row.copy()
            if 'updated_doi' not in updated_row:
                updated_row['updated_doi'] = found_doi if found_doi else original_doi
            if 'updated_error' not in updated_row:
                updated_row['updated_error'] = f"DOI found: {found_doi}" if found_doi else f"No DOI found for title '{title}'" if title else "No title available"

            updated_rows.append(updated_row)

        return pd.DataFrame(updated_rows)

    def get_unique_references(self, references_df: pd.DataFrame) -> pd.DataFrame:
        if references_df.empty:
            return pd.DataFrame()

        cache_key = id(references_df)
        if cache_key not in self._unique_references_cache:
            references_df['ref_id'] = references_df['doi'].fillna('') + '|' + references_df['title'].fillna('')
            unique_df = references_df.drop_duplicates(subset=['ref_id'], keep='first').drop(columns=['ref_id'])
            self._unique_references_cache[cache_key] = unique_df
        return self._unique_references_cache[cache_key]

    def analyze_authors_frequency(self, references_df: pd.DataFrame) -> pd.DataFrame:
        try:
            if references_df.empty:
                return pd.DataFrame()

            total_refs = len(references_df)
            unique_df = self.get_unique_references(references_df)
            total_unique = len(unique_df)
            authors_total = references_df['authors_with_initials'].str.split(',', expand=True).stack()
            authors_total = authors_total[authors_total.str.strip().isin(['Unknown', 'Error']) == False]
            author_freq_total = authors_total.value_counts().reset_index()
            author_freq_total.columns = ['author_with_initial', 'frequency_total']
            author_freq_total['percentage_total'] = round(author_freq_total['frequency_total'] / total_refs * 100, 2)
            authors_unique = unique_df['authors_with_initials'].str.split(',', expand=True).stack()
            authors_unique = authors_unique[authors_unique.str.strip().isin(['Unknown', 'Error']) == False]
            author_freq_unique = authors_unique.value_counts().reset_index()
            author_freq_unique.columns = ['author_with_initial', 'frequency_unique']
            author_freq_unique['percentage_unique'] = round(author_freq_unique['frequency_unique'] / total_unique * 100, 2)
            author_freq = author_freq_total.merge(author_freq_unique, on='author_with_initial', how='outer').fillna(0)
            return author_freq[['author_with_initial', 'frequency_total', 'percentage_total', 'frequency_unique', 'percentage_unique']].sort_values('frequency_total', ascending=False)
        except Exception as e:
            return pd.DataFrame()

    def analyze_journals_frequency(self, references_df: pd.DataFrame) -> pd.DataFrame:
        try:
            if references_df.empty:
                return pd.DataFrame()

            total_refs = len(references_df)
            unique_df = self.get_unique_references(references_df)
            total_unique = len(unique_df)
            journals_total = references_df['journal_abbreviation']
            journals_total = journals_total[journals_total.isin(['Unknown', 'Error']) == False]
            journal_freq_total = journals_total.value_counts().reset_index()
            journal_freq_total.columns = ['journal_abbreviation', 'frequency_total']
            journal_freq_total['percentage_total'] = round(journal_freq_total['frequency_total'] / total_refs * 100, 2)
            journals_unique = unique_df['journal_abbreviation']
            journals_unique = journals_unique[journals_unique.isin(['Unknown', 'Error']) == False]
            journal_freq_unique = journals_unique.value_counts().reset_index()
            journal_freq_unique.columns = ['journal_abbreviation', 'frequency_unique']
            journal_freq_unique['percentage_unique'] = round(journal_freq_unique['frequency_unique'] / total_unique * 100, 2)
            journal_freq = journal_freq_total.merge(journal_freq_unique, on='journal_abbreviation', how='outer').fillna(0)

            # Добавляем дополнительные метрики для цитирований
            journal_citation_metrics = []
            for journal in journal_freq['journal_abbreviation']:
                journal_articles = unique_df[unique_df['journal_abbreviation'] == journal]

                total_crossref_citations = journal_articles['citation_count_crossref'].sum()
                total_openalex_citations = journal_articles['citation_count_openalex'].sum()

                avg_crossref_citations = journal_articles['citation_count_crossref'].mean() if len(journal_articles) > 0 else 0
                avg_openalex_citations = journal_articles['citation_count_openalex'].mean() if len(journal_articles) > 0 else 0

                journal_citation_metrics.append({
                    'journal_abbreviation': journal,
                    'total_crossref_citations': total_crossref_citations,
                    'total_openalex_citations': total_openalex_citations,
                    'avg_crossref_citations': round(avg_crossref_citations, 2),
                    'avg_openalex_citations': round(avg_openalex_citations, 2)
                })

            journal_metrics_df = pd.DataFrame(journal_citation_metrics)
            journal_freq = journal_freq.merge(journal_metrics_df, on='journal_abbreviation', how='left').fillna(0)

            return journal_freq[['journal_abbreviation', 'frequency_total', 'percentage_total',
                               'frequency_unique', 'percentage_unique', 'total_crossref_citations',
                               'total_openalex_citations', 'avg_crossref_citations', 'avg_openalex_citations']].sort_values('frequency_total', ascending=False)
        except Exception as e:
            return pd.DataFrame()

    def analyze_affiliations_frequency(self, references_df: pd.DataFrame) -> pd.DataFrame:
        """Анализ частоты аффилиаций в references с группировкой"""
        try:
            if references_df.empty:
                return pd.DataFrame()

            total_refs = len(references_df)
            unique_df = self.get_unique_references(references_df)
            total_unique = len(unique_df)

            all_affiliations = []
            for affil_string in references_df['affiliations']:
                if pd.isna(affil_string) or affil_string in ['Unknown', 'Error', '']:
                    continue

                try:
                    affil_list = affil_string.split(';')
                    for affil in affil_list:
                        clean_affil = affil.strip()
                        if clean_affil and clean_affil not in ['Unknown', 'Error']:
                            all_affiliations.append(clean_affil)
                except Exception:
                    pass

            if not all_affiliations:
                return pd.DataFrame(columns=['affiliation', 'frequency_total', 'percentage_total', 'frequency_unique', 'percentage_unique'])

            affiliation_frequencies, grouped_organizations = self.fast_affiliation_processor.process_affiliations_list_fast(all_affiliations)

            affil_data = []
            for affil, freq in affiliation_frequencies.items():
                percentage_total = round(freq / total_refs * 100, 2) if total_refs > 0 else 0
                affil_data.append({
                    'affiliation': affil,
                    'frequency_total': freq,
                    'percentage_total': percentage_total
                })

            unique_affiliations = []
            for affil_string in unique_df['affiliations']:
                if affil_string and affil_string not in ['Unknown', 'Error']:
                    affil_list = affil_string.split(';')
                    unique_affiliations.extend([affil.strip() for affil in affil_list if affil.strip()])

            if unique_affiliations:
                unique_frequencies, _ = self.fast_affiliation_processor.process_affiliations_list_fast(unique_affiliations)

                affil_df = pd.DataFrame(affil_data)
                for affil, freq in unique_frequencies.items():
                    percentage_unique = round(freq / total_unique * 100, 2) if total_unique > 0 else 0
                    if affil in affil_df['affiliation'].values:
                        affil_df.loc[affil_df['affiliation'] == affil, 'frequency_unique'] = freq
                        affil_df.loc[affil_df['affiliation'] == affil, 'percentage_unique'] = percentage_unique
                    else:
                        affil_df = pd.concat([affil_df, pd.DataFrame([{
                            'affiliation': affil,
                            'frequency_total': 0,
                            'percentage_total': 0,
                            'frequency_unique': freq,
                            'percentage_unique': percentage_unique
                        }])], ignore_index=True)
            else:
                affil_df = pd.DataFrame(affil_data)
                affil_df['frequency_unique'] = affil_df['frequency_total']
                affil_df['percentage_unique'] = affil_df['percentage_total']

            affil_df = affil_df.fillna(0)
            return affil_df[['affiliation', 'frequency_total', 'percentage_total', 'frequency_unique', 'percentage_unique']].sort_values('frequency_total', ascending=False)

        except Exception as e:
            return pd.DataFrame()

    def analyze_countries_frequency(self, references_df: pd.DataFrame) -> pd.DataFrame:
        """Анализ частоты стран в references с разделением на отдельные страны и коллаборации"""
        try:
            if references_df.empty:
                return pd.DataFrame()

            total_refs = len(references_df)
            unique_df = self.get_unique_references(references_df)
            total_unique = len(unique_df)

            # Анализ отдельных стран
            single_countries = []
            collaborations = []

            for countries in references_df['countries']:
                if countries not in ['Unknown', 'Error']:
                    country_list = [c.strip() for c in countries.split(';')]
                    if len(country_list) == 1:
                        single_countries.extend(country_list)
                    else:
                        collaborations.append(countries)

            # Частотность отдельных стран
            single_country_counter = Counter(single_countries)
            single_country_freq = pd.DataFrame({
                'countries': list(single_country_counter.keys()),
                'type': ['single'] * len(single_country_counter),
                'frequency_total': list(single_country_counter.values())
            })
            single_country_freq['percentage_total'] = round(single_country_freq['frequency_total'] / total_refs * 100, 2)

            # Частотность коллабораций
            collaboration_counter = Counter(collaborations)
            collaboration_freq = pd.DataFrame({
                'countries': list(collaboration_counter.keys()),
                'type': ['collaboration'] * len(collaboration_counter),
                'frequency_total': list(collaboration_counter.values())
            })
            collaboration_freq['percentage_total'] = round(collaboration_freq['frequency_total'] / total_refs * 100, 2)

            # Объединяем
            country_freq_total = pd.concat([single_country_freq, collaboration_freq], ignore_index=True)

            # Аналогично для уникальных статей
            single_countries_unique = []
            collaborations_unique = []

            for countries in unique_df['countries']:
                if countries not in ['Unknown', 'Error']:
                    country_list = [c.strip() for c in countries.split(';')]
                    if len(country_list) == 1:
                        single_countries_unique.extend(country_list)
                    else:
                        collaborations_unique.append(countries)

            single_country_counter_unique = Counter(single_countries_unique)
            collaboration_counter_unique = Counter(collaborations_unique)

            single_country_freq_unique = pd.DataFrame({
                'countries': list(single_country_counter_unique.keys()),
                'frequency_unique': list(single_country_counter_unique.values())
            })
            single_country_freq_unique['percentage_unique'] = round(single_country_freq_unique['frequency_unique'] / total_unique * 100, 2)

            collaboration_freq_unique = pd.DataFrame({
                'countries': list(collaboration_counter_unique.keys()),
                'frequency_unique': list(collaboration_counter_unique.values())
            })
            collaboration_freq_unique['percentage_unique'] = round(collaboration_freq_unique['frequency_unique'] / total_unique * 100, 2)

            country_freq_unique = pd.concat([
                single_country_freq_unique.assign(type='single'),
                collaboration_freq_unique.assign(type='collaboration')
            ], ignore_index=True)

            country_freq = country_freq_total.merge(country_freq_unique, on=['countries', 'type'], how='outer').fillna(0)

            return country_freq[['countries', 'type', 'frequency_total', 'percentage_total', 'frequency_unique', 'percentage_unique']].sort_values('frequency_total', ascending=False)
        except Exception as e:
            return pd.DataFrame()

    def analyze_year_distribution(self, references_df: pd.DataFrame) -> pd.DataFrame:
        """Анализ распределения по годам для references (от новых к старым)"""
        try:
            if references_df.empty:
                return pd.DataFrame()

            total_refs = len(references_df)
            unique_df = self.get_unique_references(references_df)
            total_unique = len(unique_df)
            years_total = pd.to_numeric(references_df['year'], errors='coerce')
            years_total = years_total[years_total.notna() & years_total.between(1900, 2026)].astype(int)
            year_counts_total = years_total.value_counts().reset_index()
            year_counts_total.columns = ['year', 'frequency_total']
            year_counts_total['percentage_total'] = round(year_counts_total['frequency_total'] / total_refs * 100, 2)
            years_unique = pd.to_numeric(unique_df['year'], errors='coerce')
            years_unique = years_unique[years_unique.notna() & years_unique.between(1900, 2026)].astype(int)
            year_counts_unique = years_unique.value_counts().reset_index()
            year_counts_unique.columns = ['year', 'frequency_unique']
            year_counts_unique['percentage_unique'] = round(year_counts_unique['frequency_unique'] / total_unique * 100, 2)
            year_counts = year_counts_total.merge(year_counts_unique, on='year', how='outer').fillna(0)
            return year_counts[['year', 'frequency_total', 'percentage_total', 'frequency_unique', 'percentage_unique']].sort_values('year', ascending=False)
        except Exception as e:
            return pd.DataFrame()

    def analyze_five_year_periods(self, references_df: pd.DataFrame) -> pd.DataFrame:
        """Анализ пятилетних периодов для references (от новых к старым)"""
        try:
            if references_df.empty:
                return pd.DataFrame()

            total_refs = len(references_df)
            unique_df = self.get_unique_references(references_df)
            total_unique = len(unique_df)
            start_year = 1900
            current_year = datetime.now().year + 4
            period_starts = list(range(start_year, current_year + 1, 5))
            bins = period_starts + [period_starts[-1] + 5]
            labels = [f"{s}-{s+4}" for s in period_starts]
            years_total = pd.to_numeric(references_df['year'], errors='coerce')
            years_total = years_total[years_total.notna() & years_total.between(1900, current_year)].ast
            years_total = pd.to_numeric(references_df['year'], errors='coerce')
            years_total = years_total[years_total.notna() & years_total.between(1900, current_year)].astype(int)
            period_counts_total = pd.cut(years_total, bins=bins, labels=labels, right=False).astype(str)
            period_df_total = period_counts_total.value_counts().reset_index()
            period_df_total.columns = ['period', 'frequency_total']
            period_df_total['percentage_total'] = round(period_df_total['frequency_total'] / total_refs * 100, 2)
            period_df_total['period'] = period_df_total['period'].astype(str)
            years_unique = pd.to_numeric(unique_df['year'], errors='coerce')
            years_unique = years_unique[years_unique.notna() & years_unique.between(1900, current_year)].astype(int)
            period_counts_unique = pd.cut(years_unique, bins=bins, labels=labels, right=False).astype(str)
            period_df_unique = period_counts_unique.value_counts().reset_index()
            period_df_unique.columns = ['period', 'frequency_unique']
            period_df_unique['percentage_unique'] = round(period_df_unique['frequency_unique'] / total_unique * 100, 2)
            period_df_unique['period'] = period_df_unique['period'].astype(str)
            period_df = period_df_total.merge(period_df_unique, on='period', how='outer').fillna(0)
            return period_df[['period', 'frequency_total', 'percentage_total', 'frequency_unique', 'percentage_unique']].sort_values('period', ascending=False)
        except Exception as e:
            return pd.DataFrame()

    def find_duplicate_references(self, references_df: pd.DataFrame) -> pd.DataFrame:
        try:
            if references_df.empty:
                return pd.DataFrame()

            references_df['ref_id'] = references_df['doi'].fillna('') + '|' + references_df['title'].fillna('')
            ref_counts = references_df.groupby('ref_id')['source_doi'].nunique().reset_index()
            duplicate_ref_ids = ref_counts[ref_counts['source_doi'] > 1]['ref_id']
            if duplicate_ref_ids.empty:
                columns = list(references_df.columns) + ['frequency']
                columns.remove('ref_id')
                return pd.DataFrame(columns=columns)
            frequency_map = references_df['ref_id'].value_counts().to_dict()
            duplicates = references_df[references_df['ref_id'].isin(duplicate_ref_ids)].copy()
            duplicates = duplicates.drop_duplicates(subset=['ref_id'], keep='first')
            duplicates = duplicates[~((duplicates['doi'].isna()) & (duplicates['title'] == 'Unknown'))]
            duplicates['frequency'] = duplicates['ref_id'].map(frequency_map)
            duplicates = duplicates.drop(columns=['ref_id'])
            return duplicates.sort_values(['frequency', 'doi'], ascending=[False, True])
        except Exception as e:
            return pd.DataFrame()

    def preprocess_content_words(self, text: str) -> List[str]:
        if not text or text in ['Unknown', 'Error']:
            return []
        text = text.lower()
        text = re.sub(r'[^a-zA-Z\s-]', ' ', text)
        text = re.sub(r'\s+', ' ', text).strip()
        words = text.split()
        content_words = []
        for word in words:
            if '-' in word:
                continue
            if len(word) > 2 and word not in self.stop_words:
                stemmed_word = self.stemmer.stem(word)
                if stemmed_word not in self.scientific_stopwords_stemmed:
                    content_words.append(stemmed_word)
        return content_words

    def extract_compound_words(self, text: str) -> List[str]:
        if not text or text in ['Unknown', 'Error']:
            return []
        text = text.lower()
        compound_words = re.findall(r'\b[a-z]{2,}-[a-z]{2,}(?:-[a-z]{2,})*\b', text)
        return [word for word in compound_words if not any(part in self.stop_words for part in word.split('-'))]

    def extract_scientific_stopwords(self, text: str) -> List[str]:
        if not text or text in ['Unknown', 'Error']:
            return []
        text = text.lower()
        text = re.sub(r'[^a-zA-Z\s]', ' ', text)
        text = re.sub(r'\s+', ' ', text).strip()
        words = text.split()
        scientific_words = []
        for word in words:
            if len(word) > 2:
                stemmed_word = self.stemmer.stem(word)
                if stemmed_word in self.scientific_stopwords_stemmed:
                    for original_word in self.scientific_stopwords:
                        if self.stemmer.stem(original_word) == stemmed_word:
                            scientific_words.append(original_word)
                            break
        return scientific_words

    def analyze_titles(self, titles: List[str]) -> tuple[Counter, Counter, Counter]:
        content_words = []
        compound_words = []
        scientific_words = []
        valid_titles = [t for t in titles if t not in ['Unknown', 'Error']]
        for title in valid_titles:
            content_words.extend(self.preprocess_content_words(title))
            compound_words.extend(self.extract_compound_words(title))
            scientific_words.extend(self.extract_scientific_stopwords(title))
        return Counter(content_words), Counter(compound_words), Counter(scientific_words)

    def run_ethics_analysis(self, combined_df: pd.DataFrame, source_articles_df: pd.DataFrame) -> Dict:
        """Запуск анализа неэтичных практик"""
        try:
            st.info("🔍 Запуск анализа неэтичных практик цитирования...")
            
            # Проверяем, что данные не пустые
            if combined_df.empty or source_articles_df.empty:
                st.warning("Недостаточно данных для анализа неэтичных практик")
                return {
                    'summary': {
                        'total_findings': 0,
                        'severity_counts': {'HIGH': 0, 'MEDIUM': 0, 'LOW': 0},
                        'analysis_timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    }
                }
            
            ethics_results = self.ethics_detector.run_complete_analysis(combined_df, source_articles_df)
            
            return ethics_results
        except Exception as e:
            st.error(f"Ошибка при анализе неэтичных практик: {e}")
            return {
                'summary': {
                    'total_findings': 0,
                    'severity_counts': {'HIGH': 0, 'MEDIUM': 0, 'LOW': 0},
                    'analysis_timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                }
            }

    def display_ethics_results(self, ethics_results: Dict):
        """Отображение результатов анализа неэтичных практик"""
        st.markdown("## 🚨 Анализ неэтичных практик цитирования")
        
        summary = ethics_results.get('summary', {})
        total_findings = summary.get('total_findings', 0)
        severity_counts = summary.get('severity_counts', {})
        
        col1, col2, col3, col4 = st.columns(4)
        col1.metric("Всего находок", total_findings)
        col2.metric("Высокий риск", severity_counts.get('HIGH', 0))
        col3.metric("Средний риск", severity_counts.get('MEDIUM', 0))
        col4.metric("Низкий риск", severity_counts.get('LOW', 0))
        
        if total_findings == 0:
            st.success("✅ Неэтичные практики не обнаружены!")
            return
        
        # Отображаем находки по типам
        for practice_type, findings in ethics_results.items():
            if practice_type == 'summary' or not findings:
                continue
                
            st.markdown(f"### {self._get_practice_title(practice_type)}")
            
            for finding in findings[:10]:  # Показываем первые 10 находок каждого типа
                severity = finding.get('severity', 'LOW')
                severity_color = {
                    'HIGH': '🔴',
                    'MEDIUM': '🟡', 
                    'LOW': '🟢'
                }.get(severity, '⚪')
                
                st.markdown(f"{severity_color} **{finding.get('description', '')}**")
                
                # Дополнительная информация
                details = self._get_practice_details(finding)
                if details:
                    with st.expander("Подробности"):
                        for key, value in details.items():
                            st.write(f"**{key}:** {value}")
        
        st.warning("💡 **Примечание:** Обнаружение этих паттернов не обязательно означает нарушение, но требует дополнительной проверки.")

    def _get_practice_title(self, practice_type: str) -> str:
        """Возвращает читаемое название типа практики"""
        titles = {
            'citation_burst': 'Внезапный всплеск цитирований',
            'self_citation': 'Чрезмерное самоцитирование',
            'citation_cartels': 'Цитатные кольца',
            'newborn_citation': 'Цитирование новых статей',
            'mass_citation_author': 'Массовое цитирование автором',
            'citation_snowball': 'Цитатный снежный ком',
            'citation_templating': 'Шаблонное цитирование'
        }
        return titles.get(practice_type, practice_type)

    def _get_practice_details(self, finding: Dict) -> Dict:
        """Извлекает детали для отображения"""
        details = {}
        practice_type = finding.get('type', '')
        
        if practice_type == 'CITATION_BURST':
            details = {
                'DOI статьи': finding.get('article_doi', 'N/A'),
                'Количество цитирований': finding.get('citation_count', 'N/A'),
                'Лет с публикации': finding.get('years_since_publication', 'N/A'),
                'Соотношение цит./год': finding.get('citation_ratio', 'N/A')
            }
        elif practice_type == 'EXCESSIVE_SELF_CITATION':
            details = {
                'DOI статьи': finding.get('article_doi', 'N/A'),
                'Самоцитирования': finding.get('self_citation_count', 'N/A'),
                'Всего ссылок': finding.get('total_references', 'N/A'),
                'Процент самоцитирования': f"{finding.get('self_citation_percentage', 'N/A')}%"
            }
        elif practice_type == 'CITATION_CARTEL':
            details = {
                'Организации': ', '.join(finding.get('organizations', [])[:5]),
                'Внутренние цитирования': finding.get('internal_citations', 'N/A'),
                'Плотность цитирований': f"{finding.get('density', 'N/A'):.1%}"
            }
        elif practice_type == 'NEWBORN_CITATION':
            details = {
                'Исходная статья': finding.get('source_doi', 'N/A'),
                'Цитируемая статья': finding.get('cited_doi', 'N/A'),
                'Год цитируемой': finding.get('cited_year', 'N/A')
            }
        elif practice_type == 'MASS_CITATION_BY_AUTHOR':
            details = {
                'Автор': finding.get('author', 'N/A'),
                'Количество цитирований': finding.get('citation_count', 'N/A')
            }
        elif practice_type == 'CITATION_SNOWBALL':
            details = {
                'Статья 1': finding.get('article_doi_1', 'N/A'),
                'Статья 2': finding.get('article_doi_2', 'N/A'),
                'Общие цитирования': finding.get('common_citations', 'N/A'),
                'Схожесть': f"{finding.get('similarity', 'N/A'):.1%}"
            }
        elif practice_type == 'CITATION_TEMPLATING':
            details = {
                'Размер шаблона': finding.get('template_size', 'N/A'),
                'Статей с шаблоном': finding.get('articles_using_template', 'N/A')
            }
        
        return details

    def save_all_data_to_excel(self, combined_df: pd.DataFrame, source_articles_df: pd.DataFrame,
                         doi_list: List[str], total_references: int, unique_dois: int,
                         all_titles: List[str], ethics_results: Dict = None) -> BytesIO:
        """Сохраняет анализ references в Excel с альтметриками и этическим анализом"""
        try:
            timestamp = int(time.time())
            excel_buffer = BytesIO()

            wb = Workbook()

            # Сначала создаем вкладку Report_Summary
            ws_summary = wb.active
            ws_summary.title = 'Report_Summary'

            wb.remove(wb.active)

            try:
                unique_df = self.get_unique_references(combined_df)
            except Exception as e:
                unique_df = pd.DataFrame()

            try:
                duplicate_df = self.find_duplicate_references(combined_df)
            except Exception as e:
                duplicate_df = pd.DataFrame()

            try:
                failed_df = combined_df[combined_df['error'].notna()][['source_doi', 'position', 'doi', 'error']].copy()
                failed_df.columns = ['source_doi', 'ref_number', 'reference_doi', 'error_description']
            except Exception as e:
                failed_df = pd.DataFrame()

            stats = self.performance_monitor.get_stats()

            try:
                content_freq, compound_freq, scientific_freq = self.analyze_titles(all_titles)
            except Exception as e:
                content_freq, compound_freq, scientific_freq = Counter(), Counter(), Counter()

            try:
                total_processed = len(combined_df) if not combined_df.empty else 0

                countries_with_data = combined_df[combined_df['countries'].isin(['Unknown', 'Error']) == False]
                countries_percentage = (len(countries_with_data) / total_processed * 100) if total_processed > 0 else 0

                affiliations_with_data = combined_df[combined_df['affiliations'].isin(['Unknown', 'Error']) == False]
                affiliations_percentage = (len(affiliations_with_data) / total_processed * 100) if total_processed > 0 else 0

                altmetric_with_data = combined_df[combined_df['altmetric_score'] > 0]
                altmetric_percentage = (len(altmetric_with_data) / total_processed * 100) if total_processed > 0 else 0

            except Exception as e:
                countries_percentage = 0
                affiliations_percentage = 0
                altmetric_percentage = 0

            # Добавляем информацию об этическом анализе
            ethics_summary = ""
            if ethics_results:
                ethics_summary = f"""
ETHICS ANALYSIS SUMMARY
=======================
Total findings: {ethics_results.get('summary', {}).get('total_findings', 0)}
High severity: {ethics_results.get('summary', {}).get('severity_counts', {}).get('HIGH', 0)}
Medium severity: {ethics_results.get('summary', {}).get('severity_counts', {}).get('MEDIUM', 0)}
Low severity: {ethics_results.get('summary', {}).get('severity_counts', {}).get('LOW', 0)}

DETAILED FINDINGS:
"""
                for practice_type, findings in ethics_results.items():
                    if practice_type != 'summary' and findings:
                        ethics_summary += f"- {self._get_practice_title(practice_type)}: {len(findings)} findings\n"

            summary_content = f"""@MedvDmitry production

REFERENCES ANALYSIS REPORT

Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

ANALYSIS OVERVIEW
=================
Total source articles: {len(doi_list)}
Total references collected: {total_references}
Unique DOIs identified: {unique_dois}
Total references processed: {len(combined_df) if not combined_df.empty else 0}
Unique references: {len(unique_df) if not unique_df.empty else 0}
Successful references: {len(combined_df[combined_df['error'].isna()]) if not combined_df.empty else 0}
Failed references: {len(combined_df[combined_df['error'].notna()]) if not combined_df.empty else 0}
Unique authors: {len(self.analyze_authors_frequency(combined_df)) if not combined_df.empty else 0}
Unique journals: {len(self.analyze_journals_frequency(combined_df)) if not combined_df.empty else 0}
Unique affiliations: {len(self.analyze_affiliations_frequency(combined_df)) if not combined_df.empty else 0}
Unique countries: {len(self.analyze_countries_frequency(combined_df)) if not combined_df.empty else 0}
Duplicate references: {len(duplicate_df) if not duplicate_df.empty else 0}

DATA COMPLETENESS
=================
References with country data: {countries_percentage:.1f}%
References with affiliation data: {affiliations_percentage:.1f}%
References with altmetric data: {altmetric_percentage:.1f}%

AFFILIATION PROCESSING
======================
Affiliations normalized and grouped by organization
Similar affiliations merged together
Frequency counts reflect grouped organizations

ALTMETRIC METRICS INCLUDED
==========================
Altmetric Score: Overall attention score
Number of Mentions: Posts mentioning the article
X Mentions: Twitter/X accounts mentioning
RSS/Blogs: Blog and RSS feed mentions
Unique Accounts: Unique accounts across platforms

ETHICS ANALYSIS
===============
Advanced detection of unethical citation practices
7 different patterns analyzed
Severity levels: HIGH, MEDIUM, LOW
{ethics_summary}

PERFORMANCE STATISTICS
======================
Total processing time: {stats.get('elapsed_seconds', 0):.2f} seconds ({stats.get('elapsed_minutes', 0):.2f} minutes)
Total API requests: {stats.get('total_requests', 0)}
Requests per second: {stats.get('requests_per_second', 0):.2f}

DATA QUALITY NOTES
==================
Analysis focuses on references cited by the source articles
Combined data from Crossref and OpenAlex improves completeness
All standard statistical analyses performed (authors, journals, countries, etc.)
Error handling ensures report generation even with partial data
Affiliations normalized and grouped for consistent organization names
Altmetric metrics provide social media and online attention analysis
Ethics analysis helps identify potential citation manipulation
"""

            # Создаем вкладку Report_Summary первой
            ws_summary = wb.create_sheet('Report_Summary')
            for line in summary_content.split('\n'):
                ws_summary.append([line])

            # Удаляем колонку authors_surnames и добавляем author_count для основных таблиц
            if not source_articles_df.empty and 'authors_surnames' in source_articles_df.columns:
                source_articles_df = source_articles_df.drop(columns=['authors_surnames'])
            if not combined_df.empty and 'authors_surnames' in combined_df.columns:
                combined_df = combined_df.drop(columns=['authors_surnames'])
            if not unique_df.empty and 'authors_surnames' in unique_df.columns:
                unique_df = unique_df.drop(columns=['authors_surnames'])
            if not duplicate_df.empty and 'authors_surnames' in duplicate_df.columns:
                duplicate_df = duplicate_df.drop(columns=['authors_surnames'])

            sheets_data = [
                ('Source_Articles', source_articles_df),
                ('All_References', combined_df),
                ('All_Unique_References', unique_df),
                ('Duplicate_References', duplicate_df),
                ('Failed_References', failed_df)
            ]

            analysis_methods = [
                ('Author_Frequency', self.analyze_authors_frequency),
                ('Journal_Frequency', self.analyze_journals_frequency),
                ('Affiliation_Frequency', self.analyze_affiliations_frequency),
                ('Country_Frequency', self.analyze_countries_frequency),
                ('Year_Distribution', self.analyze_year_distribution),
                ('5_Years_Period', self.analyze_five_year_periods)
            ]

            for sheet_name, method in analysis_methods:
                try:
                    result_df = method(combined_df)
                    sheets_data.append((sheet_name, result_df))
                except Exception as e:
                    sheets_data.append((sheet_name, pd.DataFrame()))

            try:
                title_word_data = []
                for i, (word, count) in enumerate(content_freq.most_common(50), 1):
                    title_word_data.append({'Category': 'Content_Words', 'Rank': i, 'Word': word, 'Frequency': count})
                for i, (word, count) in enumerate(compound_freq.most_common(50), 1):
                    title_word_data.append({'Category': 'Compound_Words', 'Rank': i, 'Word': word, 'Frequency': count})
                for i, (word, count) in enumerate(scientific_freq.most_common(50), 1):
                    title_word_data.append({'Category': 'Scientific_Stopwords', 'Rank': i, 'Word': word, 'Frequency': count})

                title_word_df = pd.DataFrame(title_word_data)
                sheets_data.append(('Title_Word_Frequency', title_word_df))
            except Exception as e:
                sheets_data.append(('Title_Word_Frequency', pd.DataFrame()))

            # Добавляем вкладки с этическим анализом
            if ethics_results:
                for practice_type, findings in ethics_results.items():
                    if practice_type != 'summary' and findings:
                        sheet_name = f"Ethics_{practice_type.upper()}"
                        findings_df = pd.DataFrame(findings)
                        sheets_data.append((sheet_name, findings_df))

            for sheet_name, df in sheets_data:
                try:
                    if not df.empty:
                        ws = wb.create_sheet(sheet_name)
                        for r in dataframe_to_rows(df, index=False, header=True):
                            ws.append(r)
                    else:
                        ws = wb.create_sheet(sheet_name)
                        ws.append([f"No data available for {sheet_name}"])
                except Exception as e:
                    pass

            wb.save(excel_buffer)
            excel_buffer.seek(0)
            return excel_buffer

        except Exception as e:
            try:
                timestamp = int(time.time())
                excel_buffer = BytesIO()
                wb = Workbook()
                ws = wb.active
                ws.title = "Error_Report_References"
                ws.append(["ERROR REPORT - REFERENCES ANALYSIS"])
                ws.append([f"Critical error during references analysis: {str(e)}"])
                ws.append([f"Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
                ws.append(["DOIs processed:", ', '.join(doi_list)])
                wb.save(excel_buffer)
                excel_buffer.seek(0)
                return excel_buffer
            except:
                return BytesIO()

    def display_analysis_results(self, combined_df: pd.DataFrame, source_articles_df: pd.DataFrame,
                         doi_list: List[str], total_references: int, unique_dois: int,
                         all_titles: List[str], ethics_results: Dict = None) -> None:
        """Отображает результаты анализа references с альтметриками и этическим анализом"""
        try:
            st.markdown(f"{'='*80}\n**REFERENCES ANALYSIS RESULTS FOR {len(doi_list)} ARTICLES**\n{'='*80}")

            if combined_df.empty and source_articles_df.empty:
                st.error("No data available - generating error report")
                excel_buffer = self.save_all_data_to_excel(combined_df, source_articles_df, doi_list, total_references, unique_dois, all_titles, ethics_results)
                st.download_button(
                    label="Download Error Report",
                    data=excel_buffer.getvalue(),
                    file_name="error_references_report.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                return

            unique_df = self.get_unique_references(combined_df)
            successful_refs = len(combined_df[combined_df['error'].isna()])
            stats = self.performance_monitor.get_stats()

            col1, col2, col3, col4 = st.columns(4)
            col1.metric("Total references found", total_references)
            col2.metric("Unique DOIs", unique_dois)
            col3.metric("Total references processed", len(combined_df))
            col4.metric("Unique references", len(unique_df))

            col1, col2, col3, col4 = st.columns(4)
            col1.metric("Successful references", successful_refs)
            col2.metric("Failed references", len(combined_df[combined_df['error'].notna()]))
            col3.metric("Total processing time", f"{stats.get('elapsed_seconds', 0):.2f} seconds")
            col4.metric("Requests per second", f"{stats.get('requests_per_second', 0):.2f}")

            st.subheader("References per article:")
            for doi in doi_list:
                ref_count = len(combined_df[combined_df['source_doi'] == doi]) if not combined_df.empty else 0
                st.write(f"  {doi}: {ref_count} references")

            display_cols = ['source_doi', 'position', 'doi', 'title', 'authors_with_initials', 'author_count', 'year',
                          'journal_abbreviation', 'publisher', 'countries', 'citation_count_crossref',
                          'citation_count_openalex', 'altmetric_score', 'number_of_mentions']

            pd.set_option('display.max_colwidth', 25)
            pd.set_option('display.max_rows', 50)

            if not source_articles_df.empty:
                st.subheader("SOURCE ARTICLES:")
                try:
                    st.dataframe(source_articles_df[display_cols].head(10))
                except Exception as e:
                    pass

            if not unique_df.empty:
                st.subheader("UNIQUE REFERENCES:")
                try:
                    st.dataframe(unique_df[display_cols].head(10))
                except Exception as e:
                    pass

            # Отображаем результаты этического анализа
            if ethics_results:
                self.display_ethics_results(ethics_results)

            analyses = [
                ('DUPLICATE REFERENCES', self.find_duplicate_references),
                ('COUNTRIES FREQUENCY', self.analyze_countries_frequency),
                ('YEAR DISTRIBUTION', self.analyze_year_distribution),
                ('FIVE-YEAR PERIODS', self.analyze_five_year_periods),
                ('TOP 10 AUTHORS', self.analyze_authors_frequency),
                ('TOP 10 JOURNALS', self.analyze_journals_frequency),
                ('TOP 10 AFFILIATIONS', self.analyze_affiliations_frequency)
            ]

            for title, method in analyses:
                st.markdown(f"{'='*60}\n**{title}**\n{'='*60}")
                try:
                    result = method(combined_df)
                    if not result.empty:
                        if title == 'TOP 10 JOURNALS':
                            st.dataframe(result[['journal_abbreviation', 'frequency_total', 'percentage_total', 'frequency_unique', 'percentage_unique']].head(10))
                        else:
                            st.dataframe(result.head(10))
                    else:
                        st.info("No data available")
                except Exception as e:
                    st.error("Error in analysis")

            st.markdown(f"{'='*60}\n**TOP 15 TITLE WORDS**\n{'='*60}")
            try:
                content_freq, compound_freq, scientific_freq = self.analyze_titles(all_titles)
                content_df = pd.DataFrame(content_freq.most_common(15), columns=['Word', 'Frequency'])
                content_df['Category'] = 'Content_Words'
                content_df['Rank'] = range(1, len(content_df) + 1)
                compound_df = pd.DataFrame(compound_freq.most_common(15), columns=['Word', 'Frequency'])
                compound_df['Category'] = 'Compound_Words'
                compound_df['Rank'] = range(1, len(compound_df) + 1)
                scientific_df = pd.DataFrame(scientific_freq.most_common(15), columns=['Word', 'Frequency'])
                scientific_df['Category'] = 'Scientific_Stopwords'
                scientific_df['Rank'] = range(1, len(scientific_df) + 1)
                title_word_freq_df = pd.concat([content_df, compound_df, scientific_df], ignore_index=True)[['Category', 'Rank', 'Word', 'Frequency']]
                st.dataframe(title_word_freq_df)
            except Exception as e:
                st.error("Error in title analysis")

            excel_buffer = self.save_all_data_to_excel(combined_df, source_articles_df, doi_list, total_references, unique_dois, all_titles, ethics_results)
            st.download_button(
                label="Download Full Report (Excel)",
                data=excel_buffer.getvalue(),
                file_name="references_analysis_report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        except Exception as e:
            st.error(f"Critical error in display_analysis_results: {e}")
            excel_buffer = self.save_all_data_to_excel(combined_df, source_articles_df, doi_list, total_references, unique_dois, all_titles, ethics_results)
            st.download_button(
                label="Download Error Report",
                data=excel_buffer.getvalue(),
                file_name="error_references_report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

def main():
    st.title("Citation Analyzer")
    st.markdown("Analyze references and citing articles for given DOIs")

    analyzer = CitationAnalyzer()

    tab1, tab2, tab3 = st.tabs(["References Analysis", "Citing Articles Analysis", "Ethics Analysis"])

    with tab1:
        st.markdown("### Analyze References")
        st.markdown("Analyze the references cited by the input articles")
        doi_input_references = st.text_area(
            "DOIs:",
            value='10.1038/s41586-023-06924-6',
            placeholder='Enter DOIs for references analysis (e.g., 10.1010/XYZ, doi:10.1010/XYZ, https://doi.org/10.1010/XYZ, etc.) separated by any punctuation or newlines',
            height=200
        )
        
        # Добавляем опцию для включения этического анализа
        enable_ethics = st.checkbox("Включить анализ неэтичных практик", value=True, 
                                  help="Обнаруживает подозрительные паттерны цитирования")
        
        if st.button("Analyze References", type="primary"):
            with st.spinner("Processing..."):
                input_text = doi_input_references
                st.info(f"Input text: '{input_text}'")

                doi_list = analyzer.parse_doi_input(input_text)
                st.info(f"Parsed DOI list: {doi_list}")

                if not doi_list:
                    st.error("No valid DOIs provided. Please enter at least one valid DOI.")
                    st.info("Example formats:")
                    st.code("  10.1038/s41586-023-06924-6")
                    st.code("  https://doi.org/10.1038/s41586-023-06924-6")
                    st.code("  doi:10.1038/s41586-023-06924-6")
                    st.stop()

                st.success("Starting sequential processing for references analysis...")
                try:
                    combined_references_df, source_articles_df, total_references, unique_dois, all_titles = analyzer.process_doi_sequential(doi_list)
                    
                    # Запускаем этический анализ если включено
                    ethics_results = None
                    if enable_ethics and not combined_references_df.empty:
                        ethics_results = analyzer.run_ethics_analysis(combined_references_df, source_articles_df)
                    
                    analyzer.display_analysis_results(combined_references_df, source_articles_df, doi_list, total_references, unique_dois, all_titles, ethics_results)
                except Exception as e:
                    st.error(f"Critical error during processing: {e}")
                    empty_df = pd.DataFrame()
                    excel_buffer = analyzer.save_all_data_to_excel(empty_df, empty_df, doi_list, 0, 0, [])
                    st.download_button(
                        label="Download Error Report",
                        data=excel_buffer.getvalue(),
                        file_name="error_report.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    st.info("Error report generated despite processing failure.")

    with tab2:
        st.markdown("### Analyze Citing Articles")
        st.markdown("Find articles that cite the input articles (forward citations)")
        doi_input_citing = st.text_area(
            "DOIs:",
            value='10.1038/s41586-023-06924-6',
            placeholder='Enter DOIs for citing articles analysis (e.g., 10.1038/s41586-023-06924-6) separated by any punctuation or newlines',
            height=200
        )
        if st.button("Analyze Citing Articles", type="secondary"):
            with st.spinner("Processing..."):
                input_text = doi_input_citing
                st.info(f"Input text: '{input_text}'")

                doi_list = analyzer.parse_doi_input(input_text)
                st.info(f"Parsed DOI list: {doi_list}")

                if not doi_list:
                    st.error("No valid DOIs provided. Please enter at least one valid DOI.")
                    st.info("Example formats:")
                    st.code("  10.1038/s41586-023-06924-6")
                    st.code("  https://doi.org/10.1038/s41586-023-06924-6")
                    st.code("  doi:10.1038/s41586-023-06924-6")
                    st.stop()

                st.success("Starting sequential processing for citing articles analysis...")
                try:
                    citing_articles_df, citing_details_df, citing_results, all_citing_titles = analyzer.process_citing_articles_sequential(doi_list)

                    if citing_results:
                        st.markdown(f"{'='*80}\n**CITING ARTICLES ANALYSIS RESULTS**\n{'='*80}")

                        total_citation_relationships = len(citing_articles_df) if citing_articles_df is not None else 0
                        total_unique_citations = len(analyzer.get_unique_citations(citing_articles_df)) if citing_articles_df is not None and not citing_articles_df.empty else 0

                        col1, col2 = st.columns(2)
                        col1.metric("Total source articles", len(doi_list))
                        col2.metric("Total citation relationships", total_citation_relationships)
                        st.metric("Total unique citing articles", total_unique_citations)

                        st.subheader("Citations per source article:")
                        for doi, data in citing_results.items():
                            st.write(f"  {doi}: {data['count']} citations")

                        if citing_articles_df is not None and not citing_articles_df.empty:
                            st.subheader("First 10 citing articles:")
                            display_cols = ['source_doi', 'doi', 'title', 'authors_with_initials', 'author_count', 'year', 'journal_abbreviation',
                                          'citation_count_openalex', 'altmetric_score', 'number_of_mentions']
                            st.dataframe(citing_articles_df[display_cols].head(10))

                        excel_buffer = analyzer.save_citation_analysis_to_excel(citing_articles_df, citing_details_df, doi_list, citing_results, all_citing_titles)
                        st.download_button(
                            label="Download Complete Citation Analysis (Excel)",
                            data=excel_buffer.getvalue(),
                            file_name="citing_analysis_report.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                        st.success("Complete citation analysis archived and ready for download")
                    else:
                        st.warning("No citing articles found.")

                except Exception as e:
                    st.error(f"Critical error during processing: {e}")
                    empty_df = pd.DataFrame()
                    excel_buffer = analyzer.save_citation_analysis_to_excel(empty_df, empty_df, doi_list, {}, [])
                    st.download_button(
                        label="Download Error Report",
                        data=excel_buffer.getvalue(),
                        file_name="error_citing_report.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    st.info("Error report generated despite processing failure.")

    with tab3:
        st.markdown("### Ethics Analysis")
        st.markdown("Advanced detection of unethical citation practices")
        
        st.info("""
        **Обнаруживаемые практики:**
        
        🔴 **Внезапный всплеск цитирований** - статьи с аномально высоким соотношением цитирований к возрасту
        🟡 **Чрезмерное самоцитирование** - когда автор цитирует собственные работы сверх разумного предела  
        🔴 **Цитатные кольца** - группы организаций, активно цитирующих друг друга
        🟢 **Цитирование новых статей** - ссылки на очень свежие публикации
        🟡 **Массовое цитирование автором** - один автор с подозрительно большим количеством цитирований
        🟡 **Цитатный снежный ком** - статьи с очень схожими паттернами цитирования
        🔴 **Шаблонное цитирование** - одинаковые наборы цитирований в разных статьях
        """)
        
        st.warning("""
        **Важно:** Обнаружение этих паттернов не обязательно означает нарушение, 
        но указывает на практики, требующие дополнительной проверки.
        """)

if __name__ == "__main__":
    main()




